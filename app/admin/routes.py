from flask import render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash
from app.admin import bp
from app.models import User, Area, Store, Report, Region, Branch, Notification, AuditLog, ReportComment, db
from functools import wrapsfrom flask import render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash
from app.admin import bp
from app.models import User, Area, Store, Report, Region, Branch, Notification, AuditLog, ReportComment, db
from functools import wraps
from datetime import datetime, date
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz
from zoneinfo import ZoneInfo

# Timezone utilities for Egypt (Africa/Cairo) with automatic DST handling
def get_egypt_timezone():
    """Get Egypt timezone with automatic DST handling"""
    try:
        # Use zoneinfo (Python 3.9+) for better timezone handling
        return ZoneInfo("Africa/Cairo")
    except:
        # Fallback to pytz for older Python versions
        return pytz.timezone('Africa/Cairo')

def utc_to_egypt_time(utc_datetime):
    """Convert UTC datetime to Egypt local time with DST handling"""
    if utc_datetime is None:
        return None
    
    # Ensure the datetime is timezone-aware (UTC)
    if utc_datetime.tzinfo is None:
        utc_datetime = utc_datetime.replace(tzinfo=pytz.UTC)
    
    # Convert to Egypt timezone
    egypt_tz = get_egypt_timezone()
    return utc_datetime.astimezone(egypt_tz)

def egypt_time_to_utc(egypt_datetime):
    """Convert Egypt local time to UTC"""
    if egypt_datetime is None:
        return None
    
    egypt_tz = get_egypt_timezone()
    
    # If datetime is naive, assume it's in Egypt timezone
    if egypt_datetime.tzinfo is None:
        try:
            # Try zoneinfo first (Python 3.9+)
            egypt_datetime = egypt_datetime.replace(tzinfo=egypt_tz)
        except:
            # Fallback to pytz
            egypt_tz = pytz.timezone('Africa/Cairo')
            egypt_datetime = egypt_tz.localize(egypt_datetime)
    
    # Convert to UTC
    return egypt_datetime.astimezone(pytz.UTC)

def get_current_egypt_time():
    """Get current time in Egypt timezone"""
    utc_now = datetime.utcnow().replace(tzinfo=pytz.UTC)
    return utc_to_egypt_time(utc_now)

def parse_date_filter(date_str, is_end_date=False):
    """Parse date filter with proper timezone handling"""
    if not date_str:
        return None
    
    try:
        # Parse the date string (YYYY-MM-DD format from HTML date input)
        parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
        
        if is_end_date:
            # For end date, set to end of day (23:59:59) in Egypt timezone
            parsed_date = parsed_date.replace(hour=23, minute=59, second=59)
        else:
            # For start date, set to beginning of day (00:00:00) in Egypt timezone
            parsed_date = parsed_date.replace(hour=0, minute=0, second=0)
        
        # Convert Egypt local time to UTC for database query
        egypt_tz = get_egypt_timezone()
        
        try:
            # Try zoneinfo first (Python 3.9+)
            egypt_datetime = parsed_date.replace(tzinfo=egypt_tz)
        except:
            # Fallback to pytz
            egypt_tz = pytz.timezone('Africa/Cairo')
            egypt_datetime = egypt_tz.localize(parsed_date)
        
        # Convert to UTC and remove timezone info for database
        return egypt_datetime.astimezone(pytz.UTC).replace(tzinfo=None)
        
    except ValueError:
        return None

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or not session.get('is_admin', False):
            # Check if this is an API request
            if request.path.startswith('/admin/api/'):
                return jsonify({'success': False, 'error': 'Unauthorized'}), 401
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

@bp.route('/dashboard')
@admin_required
def dashboard():
    total_users = User.query.filter_by(is_admin=False).count()
    total_reports = Report.query.count()
    # Count unique branches by code (actual branches, not duplicates per supervisor)
    total_branches = db.session.query(db.func.count(db.func.distinct(Branch.code))).scalar() or 0
    # Count unique regions by name (actual regions, not duplicates per supervisor)
    total_regions = db.session.query(db.func.count(db.func.distinct(Region.name))).scalar() or 0
    
    # Recent reports with timezone conversion
    recent_reports_raw = Report.query.order_by(Report.created_at.desc()).limit(5).all()
    recent_reports = []
    for report in recent_reports_raw:
        # Convert UTC to Egypt time
        report.submitted_time_egypt = utc_to_egypt_time(report.created_at)
        recent_reports.append(report)
    
    return render_template('admin/dashboard.html', 
                         total_users=total_users,
                         total_reports=total_reports,
                         total_branches=total_branches,
                         total_regions=total_regions,
                         recent_reports=recent_reports,
                         utc_to_egypt_time=utc_to_egypt_time)

@bp.route('/api/dashboard/stats', methods=['GET'])
@admin_required
def get_dashboard_stats():
    """API endpoint to get real-time dashboard statistics"""
    try:
        # Get fresh counts from database using COUNT queries
        total_users = db.session.query(db.func.count(User.id)).filter(User.is_admin == False).scalar() or 0
        total_reports = db.session.query(db.func.count(Report.id)).scalar() or 0
        # Count unique branches by code (actual branches, not duplicates per supervisor)
        total_branches = db.session.query(db.func.count(db.func.distinct(Branch.code))).scalar() or 0
        # Count unique regions by name (actual regions, not duplicates per supervisor)
        total_regions = db.session.query(db.func.count(db.func.distinct(Region.name))).scalar() or 0
        
        return jsonify({
            'success': True,
            'stats': {
                'total_users': total_users,
                'total_reports': total_reports,
                'total_branches': total_branches,
                'total_regions': total_regions
            }
        })
    except Exception as e:
        print(f"Error in get_dashboard_stats: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================================
# NEW CLEAN USERS MANAGEMENT SYSTEM
# ============================================================================

@bp.route('/users')
@admin_required
def manage_users():
    """صفحة إدارة المستخدمين الجديدة والنظيفة"""
    return render_template('admin/users_management.html')

@bp.route('/admin-users')
@admin_required
def admin_users():
    """صفحة إدارة المستخدمين الإداريين"""
    return render_template('admin/admin_users_management.html')

# ============================================================================
# USERS API - Clean and Simple
# ============================================================================

@bp.route('/api/users', methods=['GET'])
@admin_required
def api_get_users():
    """Get all non-admin users with their regions and branches"""
    try:
        users = User.query.filter_by(is_admin=False).all()
        users_data = []
        
        for user in users:
            # Get user's owned regions and branches
            user_regions = Region.query.filter_by(owner_user_id=user.id).all()
            regions_data = []
            
            for region in user_regions:
                region_branches = Branch.query.filter_by(
                    owner_user_id=user.id, 
                    region_id=region.id
                ).all()
                
                branches_data = [{
                    'id': branch.id,
                    'name': branch.name,
                    'code': branch.code,
                    'governorate': branch.governorate
                } for branch in region_branches]
                
                regions_data.append({
                    'id': region.id,
                    'name': region.name,
                    'branches': branches_data
                })
            
            users_data.append({
                'id': user.id,
                'spvr_name': user.employee_name,
                'spvr_code': user.employee_code,
                'username': user.username,
                'created_at': user.created_at.isoformat(),
                'regions': regions_data
            })
        
        return jsonify(users_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    """Create a new user with regions and branches"""
    try:
        data = request.get_json()
        
        # Validate required fields
        spvr_name = data.get('spvr_name', '').strip()
        spvr_code = data.get('spvr_code', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        regions_data = data.get('regions', [])
        
        if not all([spvr_name, spvr_code, username, password]):
            return jsonify({
                'success': False, 
                'message': 'All required fields must be filled'
            }), 400
        
        # Check for existing username or employee code
        if User.query.filter_by(username=username).first():
            return jsonify({
                'success': False, 
                'message': 'Username already exists'
            }), 409
        
        if User.query.filter_by(employee_code=spvr_code).first():
            return jsonify({
                'success': False, 
                'message': 'SPVR Code already exists'
            }), 409
        
        # Create user
        user = User(
            employee_name=spvr_name,
            employee_code=spvr_code,
            username=username,
            password_hash=generate_password_hash(password),
            is_admin=False
        )
        
        db.session.add(user)
        db.session.flush()  # Get user ID
        
        # Create regions and branches
        created_regions = []
        created_branches = []
        
        for region_data in regions_data:
            region_name = region_data.get('name', '').strip()
            
            if not region_name:
                continue
            
            # Create region
            region = Region(
                name=region_name,
                owner_user_id=user.id
            )
            db.session.add(region)
            db.session.flush()  # Get region ID
            created_regions.append(region)
            
            # Create branches for this region
            branches_data = region_data.get('branches', [])
            for branch_data in branches_data:
                branch_name = branch_data.get('name', '').strip()
                branch_code = branch_data.get('code', '').strip()
                branch_governorate = branch_data.get('governorate', '').strip()
                
                if not branch_name or not branch_code:
                    continue
                
                # Check for duplicate branch codes for this user
                existing_branch = Branch.query.filter_by(
                    code=branch_code,
                    owner_user_id=user.id
                ).first()
                
                if existing_branch:
                    raise ValueError(f'Branch code "{branch_code}" already exists for this user')
                
                # Create branch
                branch = Branch(
                    name=branch_name,
                    code=branch_code,
                    governorate=branch_governorate if branch_governorate else None,
                    region_id=region.id,
                    owner_user_id=user.id
                )
                db.session.add(branch)
                created_branches.append(branch)
        
        # Commit all changes
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'User created successfully',
            'user_id': user.id,
            'regions_created': len(created_regions),
            'branches_created': len(created_branches)
        }), 201
        
    except ValueError as ve:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(ve)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users/<int:user_id>', methods=['GET'])
@admin_required
def api_get_user(user_id):
    """Get a specific user with regions and branches"""
    try:
        user = User.query.filter_by(id=user_id, is_admin=False).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get user's owned regions and branches
        user_regions = Region.query.filter_by(owner_user_id=user.id).all()
        regions_data = []
        
        for region in user_regions:
            region_branches = Branch.query.filter_by(
                owner_user_id=user.id, 
                region_id=region.id
            ).all()
            
            branches_data = [{
                'id': branch.id,
                'name': branch.name,
                'code': branch.code,
                'governorate': branch.governorate
            } for branch in region_branches]
            
            regions_data.append({
                'id': region.id,
                'name': region.name,
                'branches': branches_data
            })
        
        user_data = {
            'id': user.id,
            'spvr_name': user.employee_name,
            'spvr_code': user.employee_code,
            'username': user.username,
            'created_at': user.created_at.isoformat(),
            'regions': regions_data
        }
        
        return jsonify(user_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def api_update_user(user_id):
    """Update a user with regions and branches"""
    try:
        user = User.query.filter_by(id=user_id, is_admin=False).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        data = request.get_json()
        
        # Validate required fields
        spvr_name = data.get('spvr_name', '').strip()
        spvr_code = data.get('spvr_code', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        regions_data = data.get('regions', [])
        
        if not all([spvr_name, spvr_code, username]):
            return jsonify({
                'success': False, 
                'message': 'Name, code, and username are required'
            }), 400
        
        # Check for existing username or employee code (excluding current user)
        existing_user = User.query.filter(
            User.username == username,
            User.id != user_id
        ).first()
        if existing_user:
            return jsonify({
                'success': False, 
                'message': 'Username already exists'
            }), 409
        
        existing_user = User.query.filter(
            User.employee_code == spvr_code,
            User.id != user_id
        ).first()
        if existing_user:
            return jsonify({
                'success': False, 
                'message': 'SPVR Code already exists'
            }), 409
        
        # Update user basic info
        user.employee_name = spvr_name
        user.employee_code = spvr_code
        user.username = username
        
        if password:
            user.password_hash = generate_password_hash(password)
        
        # Delete existing regions and branches for this user
        existing_branches = Branch.query.filter_by(owner_user_id=user.id).all()
        for branch in existing_branches:
            db.session.delete(branch)
        
        existing_regions = Region.query.filter_by(owner_user_id=user.id).all()
        for region in existing_regions:
            db.session.delete(region)
        
        # Flush the deletions to ensure they're processed before creating new ones
        db.session.flush()
        
        # Create new regions and branches
        created_regions = []
        created_branches = []
        
        for region_data in regions_data:
            region_name = region_data.get('name', '').strip()
            
            if not region_name:
                continue
            
            # Create region
            region = Region(
                name=region_name,
                owner_user_id=user.id
            )
            db.session.add(region)
            db.session.flush()  # Get region ID
            created_regions.append(region)
            
            # Create branches for this region
            branches_data = region_data.get('branches', [])
            for branch_data in branches_data:
                branch_name = branch_data.get('name', '').strip()
                branch_code = branch_data.get('code', '').strip()
                branch_governorate = branch_data.get('governorate', '').strip()
                
                if not branch_name or not branch_code:
                    continue
                
                # Create branch
                branch = Branch(
                    name=branch_name,
                    code=branch_code,
                    governorate=branch_governorate if branch_governorate else None,
                    region_id=region.id,
                    owner_user_id=user.id
                )
                db.session.add(branch)
                created_branches.append(branch)
        
        # Commit all changes
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'User updated successfully',
            'regions_created': len(created_regions),
            'branches_created': len(created_branches)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_delete_user(user_id):
    """Delete a user and all their owned regions/branches"""
    try:
        user = User.query.filter_by(id=user_id, is_admin=False).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get request data to check for force delete or reassign options
        data = request.get_json(silent=True) or {}
        force_delete = data.get('force_delete', False)
        reassign_to_user_id = data.get('reassign_to_user_id')
        
        # Convert reassign_to_user_id to integer if it exists
        if reassign_to_user_id is not None:
            try:
                reassign_to_user_id = int(reassign_to_user_id)
            except (ValueError, TypeError):
                return jsonify({
                    'success': False, 
                    'message': 'Invalid user ID for reassignment'
                }), 400
        
        # Check if user has reports
        reports_count = len(user.reports)
        if reports_count > 0:
            if force_delete:
                # Delete all reports associated with this user
                for report in user.reports:
                    db.session.delete(report)
            else:
                # Return info about reports and force delete option
                return jsonify({
                    'success': False, 
                    'message': f'Cannot delete user. User has {reports_count} report(s). Use force delete to remove user and all reports.',
                    'reports_count': reports_count,
                    'force_delete_required': True
                }), 400
        
        # Delete user's branches first (due to foreign key constraints)
        user_branches = Branch.query.filter_by(owner_user_id=user.id).all()
        for branch in user_branches:
            db.session.delete(branch)
        
        # Delete user's regions
        user_regions = Region.query.filter_by(owner_user_id=user.id).all()
        for region in user_regions:
            db.session.delete(region)
        
        # Delete user's vacations (to avoid NOT NULL constraint error)
        from app.models import Vacation
        user_vacations = Vacation.query.filter_by(user_id=user.id).all()
        for vacation in user_vacations:
            db.session.delete(vacation)
        
        # Delete user's notifications (to avoid NOT NULL constraint error)
        user_notifications = Notification.query.filter_by(user_id=user.id).all()
        for notification in user_notifications:
            db.session.delete(notification)
        
        # Delete user's comments (to avoid NOT NULL constraint error)
        user_comments = ReportComment.query.filter_by(user_id=user.id).all()
        for comment in user_comments:
            db.session.delete(comment)
        
        # Delete user's audit logs (optional, but good for cleanup)
        user_audit_logs = AuditLog.query.filter_by(user_id=user.id).all()
        for audit_log in user_audit_logs:
            db.session.delete(audit_log)
        
        # Clear any associations (backward compatibility)
        user.assigned_areas.clear()
        user.assigned_stores.clear()
        user.assigned_regions.clear()
        user.assigned_branches.clear()
        
        # Delete user
        db.session.delete(user)
        db.session.commit()
        
        action_taken = "deleted"
        if reports_count > 0 and force_delete:
            action_taken = f"deleted along with {reports_count} reports"
        
        return jsonify({
            'success': True,
            'message': f'User {action_taken} successfully',
            'regions_deleted': len(user_regions),
            'branches_deleted': len(user_branches),
            'vacations_deleted': len(user_vacations),
            'reports_deleted': reports_count if force_delete else 0
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users/<int:user_id>/reports', methods=['GET'])
@admin_required
def api_get_user_reports(user_id):
    """Get reports for a specific user"""
    try:
        user = User.query.filter_by(id=user_id, is_admin=False).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        reports_data = []
        for report in user.reports:
            # Convert UTC times to Egypt local time for display
            report_date_local = utc_to_egypt_time(report.report_date)
            created_at_local = utc_to_egypt_time(report.created_at)
            
            reports_data.append({
                'id': report.id,
                'store_name': report.store.name,
                'store_code': report.store.code,
                'area': report.area.name,
                'report_date': report_date_local.strftime('%Y-%m-%d') if report_date_local else '',
                'created_at': created_at_local.strftime('%Y-%m-%d %H:%M') if created_at_local else ''
            })
        
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'name': user.employee_name,
                'code': user.employee_code
            },
            'reports': reports_data,
            'reports_count': len(reports_data)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/reports/<int:report_id>', methods=['DELETE'])
@admin_required
def api_delete_report(report_id):
    """Delete a specific report"""
    try:
        report = Report.query.get(report_id)
        if not report:
            return jsonify({'success': False, 'message': 'Report not found'}), 404
        
        # Store report info for response
        report_info = {
            'id': report.id,
            'employee_name': report.employee.employee_name,
            'store_name': report.store.name,
            'report_date': report.report_date.strftime('%Y-%m-%d') if report.report_date else 'N/A'
        }
        
        # Delete the report
        db.session.delete(report)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Report deleted successfully',
            'deleted_report': report_info
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/reports/delete-all', methods=['DELETE'])
@admin_required
def api_delete_all_reports():
    """Delete all reports (with optional filters)"""
    try:
        # Get filter parameters (same as in api_reports)
        employee_name = request.args.get('employee_name', '')
        employee_code = request.args.get('employee_code', '')
        store_name = request.args.get('store_name', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        print(f"🗑️ Delete all reports request received")
        print(f"   Filters: employee_name={employee_name}, employee_code={employee_code}, store_name={store_name}")
        print(f"   Date range: {start_date} to {end_date}")
        
        # Build query with same filters as api_reports
        query = db.session.query(Report).join(User).join(Store).join(Area)
        
        if employee_name:
            query = query.filter(User.employee_name.contains(employee_name))
        if employee_code:
            query = query.filter(User.employee_code.contains(employee_code))
        if store_name:
            query = query.filter(Store.name.contains(store_name))
        
        # Date filtering
        if start_date:
            start_datetime = parse_date_filter(start_date, is_end_date=False)
            if start_datetime:
                query = query.filter(Report.report_date >= start_datetime)
        
        if end_date:
            end_datetime = parse_date_filter(end_date, is_end_date=True)
            if end_datetime:
                query = query.filter(Report.report_date <= end_datetime)
        
        # Get all matching reports
        reports_to_delete = query.all()
        deleted_count = len(reports_to_delete)
        
        print(f"   Found {deleted_count} reports to delete")
        
        if deleted_count == 0:
            return jsonify({
                'success': False,
                'message': 'No reports found matching the criteria',
                'deleted_count': 0
            }), 404
        
        # Delete all matching reports
        for report in reports_to_delete:
            db.session.delete(report)
        
        db.session.commit()
        
        print(f"✅ Successfully deleted {deleted_count} reports")
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted {deleted_count} report(s)',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        print(f"❌ Error deleting all reports: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================================
# ADMIN USERS MANAGEMENT API
# ============================================================================

@bp.route('/api/admin-users', methods=['GET'])
@admin_required
def api_get_admin_users():
    """Get all admin users"""
    try:
        admins = User.query.filter_by(is_admin=True).all()
        admins_data = []
        
        for admin in admins:
            # Determine if this is the primary admin (first admin created)
            is_primary = admin.id == 1
            
            admins_data.append({
                'id': admin.id,
                'employee_name': admin.employee_name,
                'employee_code': admin.employee_code,
                'username': admin.username,
                'is_primary_admin': is_primary,
                'created_at': admin.created_at.isoformat()
            })
        
        return jsonify(admins_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/admin-users', methods=['POST'])
@admin_required
def api_create_admin_user():
    """Create a new admin user"""
    try:
        data = request.get_json()
        
        # Validate required fields
        name = data.get('employee_name', '').strip()
        employee_code = data.get('employee_code', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not all([name, username, password]):
            return jsonify({
                'success': False, 
                'message': 'Name, username, and password are required'
            }), 400
        
        if len(password) < 8:
            return jsonify({
                'success': False, 
                'message': 'Password must be at least 8 characters'
            }), 400
        
        # Check for existing username
        if User.query.filter_by(username=username).first():
            return jsonify({
                'success': False, 
                'message': 'Username already exists'
            }), 409
        
        # Use provided employee code or generate one
        if not employee_code:
            import uuid
            employee_code = f"ADMIN_{uuid.uuid4().hex[:8].upper()}"
        
        # Create admin user
        admin = User(
            employee_name=name,
            employee_code=employee_code,
            username=username,
            password_hash=generate_password_hash(password),
            is_admin=True
        )
        
        db.session.add(admin)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Administrator created successfully',
            'admin_id': admin.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/admin-users/<int:admin_id>', methods=['PUT'])
@admin_required
def api_update_admin_user(admin_id):
    """Update an admin user"""
    try:
        admin = User.query.filter_by(id=admin_id, is_admin=True).first()
        if not admin:
            return jsonify({'success': False, 'message': 'Admin user not found'}), 404
        
        data = request.get_json()
        
        # Validate required fields
        name = data.get('employee_name', '').strip()
        employee_code = data.get('employee_code', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not all([name, username]):
            return jsonify({
                'success': False, 
                'message': 'Name and username are required'
            }), 400
        
        if password and len(password) < 8:
            return jsonify({
                'success': False, 
                'message': 'Password must be at least 8 characters'
            }), 400
        
        # Check for existing username (excluding current admin)
        existing_user = User.query.filter(
            User.username == username,
            User.id != admin_id
        ).first()
        if existing_user:
            return jsonify({
                'success': False, 
                'message': 'Username already exists'
            }), 409
        
        # Update admin
        admin.employee_name = name
        if employee_code:
            admin.employee_code = employee_code
        admin.username = username
        
        if password:
            admin.password_hash = generate_password_hash(password)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Administrator updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/admin-users/<int:admin_id>', methods=['DELETE'])
@admin_required
def api_delete_admin_user(admin_id):
    """Delete an admin user"""
    try:
        admin = User.query.filter_by(id=admin_id, is_admin=True).first()
        if not admin:
            return jsonify({'success': False, 'message': 'Admin user not found'}), 404
        
        # Check if this is the primary admin
        if admin.id == 1:
            return jsonify({
                'success': False, 
                'message': 'Cannot delete the primary administrator'
            }), 400
        
        # Check if this would be the last admin
        admin_count = User.query.filter_by(is_admin=True).count()
        if admin_count <= 1:
            return jsonify({
                'success': False, 
                'message': 'Cannot delete the last administrator account'
            }), 400
        
        # Delete admin's notifications (to avoid NOT NULL constraint error)
        admin_notifications = Notification.query.filter_by(user_id=admin.id).all()
        for notification in admin_notifications:
            db.session.delete(notification)
        
        # Delete admin's comments (to avoid NOT NULL constraint error)
        admin_comments = ReportComment.query.filter_by(user_id=admin.id).all()
        for comment in admin_comments:
            db.session.delete(comment)
        
        # Delete admin's audit logs (optional, but good for cleanup)
        admin_audit_logs = AuditLog.query.filter_by(user_id=admin.id).all()
        for audit_log in admin_audit_logs:
            db.session.delete(audit_log)
        
        # Delete admin
        db.session.delete(admin)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Administrator deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================================
# REPORTS MANAGEMENT (keeping existing functionality)
# ============================================================================

@bp.route('/reports')
@admin_required
def reports():
    return render_template('admin/reports.html')

@bp.route('/reports/<int:report_id>/view')
@admin_required
def view_report_detail(report_id):
    """عرض تفاصيل تقرير معين"""
    report = Report.query.get_or_404(report_id)
    
    # Mark report as read by admin
    if not report.is_read:
        report.is_read = True
        db.session.commit()
    
    return render_template('admin/view_report_detail.html', report=report)

@bp.route('/api/reports')
@admin_required
def api_reports():
    # Get filter parameters
    employee_name = request.args.get('employee_name', '')
    employee_code = request.args.get('employee_code', '')
    store_name = request.args.get('store_name', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build query
    query = db.session.query(Report).join(User).join(Store)
    
    if employee_name:
        query = query.filter(User.employee_name.contains(employee_name))
    if employee_code:
        query = query.filter(User.employee_code.contains(employee_code))
    if store_name:
        query = query.filter(Store.name.contains(store_name))
    
    # Fixed date filtering with proper timezone handling
    if start_date:
        start_datetime = parse_date_filter(start_date, is_end_date=False)
        if start_datetime:
            query = query.filter(Report.report_date >= start_datetime)
    
    if end_date:
        end_datetime = parse_date_filter(end_date, is_end_date=True)
        if end_datetime:
            query = query.filter(Report.report_date <= end_datetime)
    
    reports = query.order_by(Report.created_at.desc()).all()
    
    reports_data = []
    for report in reports:
        # Convert UTC times to Egypt local time for display
        report_date_local = utc_to_egypt_time(report.report_date)
        created_at_local = utc_to_egypt_time(report.created_at)
        
        # Count comments
        from app.models import ReportComment
        comments_count = ReportComment.query.filter_by(report_id=report.id).count()
        
        reports_data.append({
            'id': report.id,
            'employee_name': report.employee.employee_name,
            'employee_code': report.employee.employee_code,
            'store_name': report.store.name,
            'store_code': report.store.code,
            'area': report.area.name,
            'report_date': report_date_local.strftime('%Y-%m-%d %H:%M') if report_date_local else '',
            'created_at': created_at_local.strftime('%Y-%m-%d %H:%M') if created_at_local else '',
            'status': report.status if hasattr(report, 'status') else 'new',
            'is_read': report.is_read if hasattr(report, 'is_read') else False,
            'comments_count': comments_count
        })
    
    return jsonify(reports_data)

@bp.route('/preview_export')
@admin_required
def preview_export():
    """Preview export data before downloading"""
    # Get same filters as API
    employee_name = request.args.get('employee_name', '')
    employee_code = request.args.get('employee_code', '')
    store_name = request.args.get('store_name', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build query
    query = db.session.query(Report).join(User).join(Store).join(Area)
    
    if employee_name:
        query = query.filter(User.employee_name.contains(employee_name))
    if employee_code:
        query = query.filter(User.employee_code.contains(employee_code))
    if store_name:
        query = query.filter(Store.name.contains(store_name))
    
    # Fixed date filtering with proper timezone handling
    if start_date:
        start_datetime = parse_date_filter(start_date, is_end_date=False)
        if start_datetime:
            query = query.filter(Report.report_date >= start_datetime)
    
    if end_date:
        end_datetime = parse_date_filter(end_date, is_end_date=True)
        if end_datetime:
            query = query.filter(Report.report_date <= end_datetime)
    
    reports_raw = query.order_by(Report.created_at.desc()).limit(100).all()  # Limit for preview
    
    # Add governorate to each report
    reports = []
    for report in reports_raw:
        # Get governorate from store, area, or user's branch assignments
        governorate = ''
        
        # First try to get from store
        if hasattr(report, 'store') and hasattr(report.store, 'governorate') and report.store.governorate:
            governorate = report.store.governorate
        # Then try to get from area
        elif hasattr(report, 'area') and hasattr(report.area, 'governorate') and report.area.governorate:
            governorate = report.area.governorate
        # Finally try to get from user's branch assignments
        else:
            # Get the first branch with a governorate for this user
            user_branches = Branch.query.filter_by(owner_user_id=report.user_id).all()
            for branch in user_branches:
                if branch.governorate:
                    governorate = branch.governorate
                    break
        
        # Add governorate as attribute
        report.governorate_display = governorate
        reports.append(report)
    
    return render_template('admin/export_preview.html', reports=reports)

@bp.route('/export')
@admin_required
def export_reports():
    """Export reports with professional formatting and separate sheets per SPVR"""
    # Get same filters as API
    employee_name = request.args.get('employee_name', '')
    employee_code = request.args.get('employee_code', '')
    store_name = request.args.get('store_name', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build query
    query = db.session.query(Report).join(User).join(Store).join(Area)
    
    if employee_name:
        query = query.filter(User.employee_name.contains(employee_name))
    if employee_code:
        query = query.filter(User.employee_code.contains(employee_code))
    if store_name:
        query = query.filter(Store.name.contains(store_name))
    
    # Fixed date filtering with proper timezone handling
    if start_date:
        start_datetime = parse_date_filter(start_date, is_end_date=False)
        if start_datetime:
            query = query.filter(Report.report_date >= start_datetime)
    
    if end_date:
        end_datetime = parse_date_filter(end_date, is_end_date=True)
        if end_datetime:
            query = query.filter(Report.report_date <= end_datetime)
    
    reports = query.order_by(Report.created_at.desc()).all()
    
    # Generate filename with current Egypt local date and time
    current_egypt_time = get_current_egypt_time()
    filename = f'Report_{current_egypt_time.strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    if not reports:
        # Even if no reports, create summary sheet to show vacation status
        output = io.BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet with empty reports data
        summary_ws = wb.create_sheet(title="Reports Summary")
        _create_summary_sheet(summary_ws, {}, start_date, end_date)  # Empty spvr_reports dict
        
        # Create a "No Reports" sheet for information
        no_reports_ws = wb.create_sheet(title="No Reports Found")
        no_reports_ws.cell(row=1, column=1, value="No reports found for the selected criteria")
        no_reports_ws.cell(row=2, column=1, value=f"Date range: {start_date or 'All'} to {end_date or 'All'}")
        no_reports_ws.cell(row=3, column=1, value="Check the 'Reports Summary' sheet for employee vacation status")
        
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    # Group reports by SPVR
    spvr_reports = {}
    for report in reports:
        spvr_key = f"{report.employee.employee_code}_{report.employee.employee_name}"
        if spvr_key not in spvr_reports:
            spvr_reports[spvr_key] = []
        spvr_reports[spvr_key].append(report)
    
    # Create Excel file with professional formatting
    output = io.BytesIO()
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create summary sheet first
    summary_ws = wb.create_sheet(title="Reports Summary")
    _create_summary_sheet(summary_ws, spvr_reports, start_date, end_date)
    
    # Create a sheet for each SPVR with enhanced naming
    used_sheet_names = set()
    used_sheet_names.add("Reports Summary")  # Reserve summary sheet name
    
    for spvr_key, spvr_report_list in spvr_reports.items():
        spvr_code, spvr_name = spvr_key.split('_', 1)
        
        # Create enhanced sheet name with code and name
        sheet_name = _create_sheet_name(spvr_code, spvr_name, len(spvr_reports) > 1, used_sheet_names)
        used_sheet_names.add(sheet_name)
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Apply enhanced professional formatting to this sheet
        _format_excel_sheet_enhanced(ws, spvr_report_list, spvr_name, spvr_code)
    
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _create_summary_sheet(ws, spvr_reports, start_date, end_date):
    """Create summary sheet with employee statistics and missing employees in red"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models import User, Store, Branch, Vacation
    from datetime import datetime, date
    
    # Define colors and styles
    header_color = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue header
    summary_color = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Light green for active employees
    missing_color = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # Light red for missing employees
    vacation_color = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow for vacation
    
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10, bold=False, color='000000')
    missing_font = Font(name='Calibri', size=10, bold=True, color='CC0000')  # Red bold for missing
    vacation_font = Font(name='Calibri', size=10, bold=True, color='FF8C00')  # Orange bold for vacation
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Reports Summary - From {start_date or 'Beginning'} To {end_date or 'End'}"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='000000')
    title_cell.alignment = center_alignment
    
    # Headers
    headers = ['Employee Code', 'Employee Name', 'Stores Count', 'Reports Count', 'Last Report', 'Status']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_color
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Get all employees (non-admin users)
    all_employees = User.query.filter_by(is_admin=False).all()
    
    # Parse date range for vacation checking
    vacation_start_date = None
    vacation_end_date = None
    
    if start_date:
        try:
            vacation_start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        except:
            # If parsing fails, get earliest vacation date or use 30 days ago
            earliest_vacation = Vacation.query.order_by(Vacation.vacation_date.asc()).first()
            if earliest_vacation:
                vacation_start_date = earliest_vacation.vacation_date
            else:
                vacation_start_date = date.today() - timedelta(days=30)
    else:
        # If no start date provided, get earliest vacation date or use 30 days ago
        earliest_vacation = Vacation.query.order_by(Vacation.vacation_date.asc()).first()
        if earliest_vacation:
            vacation_start_date = earliest_vacation.vacation_date
        else:
            vacation_start_date = date.today() - timedelta(days=30)
        
    if end_date:
        try:
            vacation_end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except:
            # If parsing fails, get latest vacation date or use today
            latest_vacation = Vacation.query.order_by(Vacation.vacation_date.desc()).first()
            if latest_vacation:
                vacation_end_date = latest_vacation.vacation_date
            else:
                vacation_end_date = date.today()
    else:
        # If no end date provided, get latest vacation date or use today
        latest_vacation = Vacation.query.order_by(Vacation.vacation_date.desc()).first()
        if latest_vacation:
            vacation_end_date = latest_vacation.vacation_date
        else:
            vacation_end_date = date.today()
    
    print(f"🏖️ Checking vacations from {vacation_start_date} to {vacation_end_date}")
    
    # Create employee statistics
    employee_stats = {}
    
    # Process employees who submitted reports
    for spvr_key, reports in spvr_reports.items():
        spvr_code, spvr_name = spvr_key.split('_', 1)
        
        # Get unique stores reported by this employee
        unique_stores = set()
        last_report_date = None
        
        for report in reports:
            unique_stores.add(report.store.id)
            if not last_report_date or report.report_date > last_report_date:
                last_report_date = report.report_date
        
        employee_stats[spvr_code] = {
            'name': spvr_name,
            'code': spvr_code,
            'stores_count': len(unique_stores),
            'reports_count': len(reports),
            'last_report': last_report_date,
            'has_reports': True,
            'has_vacation': False
        }
    
    # Add employees who didn't submit reports
    for employee in all_employees:
        if employee.employee_code not in employee_stats:
            employee_stats[employee.employee_code] = {
                'name': employee.employee_name,
                'code': employee.employee_code,
                'stores_count': 0,
                'reports_count': 0,
                'last_report': None,
                'has_reports': False,
                'has_vacation': False
            }
    
    # Now check vacations for ALL employees (both with and without reports)
    for employee in all_employees:
        if employee.employee_code in employee_stats:
            # Check if employee has vacation in the date range - GET ALL vacations, not just first
            vacations = Vacation.query.filter(
                Vacation.user_id == employee.id,
                Vacation.vacation_date >= vacation_start_date,
                Vacation.vacation_date <= vacation_end_date
            ).all()
            
            has_vacation = len(vacations) > 0
            employee_stats[employee.employee_code]['has_vacation'] = has_vacation
            
            if has_vacation:
                vacation_dates = [v.vacation_date.strftime('%Y-%m-%d') for v in vacations]
                print(f"🏖️ Found {len(vacations)} vacation(s) for {employee.employee_name} ({employee.employee_code}): {', '.join(vacation_dates)}")
            else:
                print(f"💼 No vacation found for {employee.employee_name} ({employee.employee_code})")
    
    # Sort employees by name
    sorted_employees = sorted(employee_stats.values(), key=lambda x: x['name'])
    
    # Write data
    row_idx = 4
    for emp in sorted_employees:
        # Employee code
        cell = ws.cell(row=row_idx, column=1)
        cell.value = emp['code']
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Employee name
        cell = ws.cell(row=row_idx, column=2)
        cell.value = emp['name']
        cell.border = thin_border
        cell.alignment = left_alignment
        
        # Stores count
        cell = ws.cell(row=row_idx, column=3)
        cell.value = emp['stores_count']
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Reports count
        cell = ws.cell(row=row_idx, column=4)
        cell.value = emp['reports_count']
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Last report date
        cell = ws.cell(row=row_idx, column=5)
        if emp['last_report']:
            last_report_local = utc_to_egypt_time(emp['last_report'])
            cell.value = last_report_local.strftime('%Y-%m-%d') if last_report_local else ''
        else:
            cell.value = 'No Reports'
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Status
        cell = ws.cell(row=row_idx, column=6)
        if emp['has_vacation']:
            cell.value = 'On Vacation'
            cell.fill = vacation_color
            cell.font = vacation_font
        elif emp['has_reports']:
            cell.value = 'Active'
            cell.fill = summary_color
            cell.font = data_font
        else:
            cell.value = 'No Reports'
            cell.fill = missing_color
            cell.font = missing_font
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Apply row formatting
        for col in range(1, 7):
            row_cell = ws.cell(row=row_idx, column=col)
            if emp['has_vacation']:
                if col != 6:  # Don't override status cell
                    row_cell.fill = vacation_color
                    row_cell.font = vacation_font
            elif not emp['has_reports']:
                if col != 6:  # Don't override status cell
                    row_cell.fill = missing_color
                    row_cell.font = missing_font
            else:
                if col != 6:  # Don't override status cell
                    row_cell.font = data_font
        
        row_idx += 1
    
    # Set column widths
    column_widths = [15, 25, 15, 15, 15, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add summary statistics at the bottom
    summary_row = row_idx + 2
    
    total_employees = len(sorted_employees)
    active_employees = sum(1 for emp in sorted_employees if emp['has_reports'])
    vacation_employees = sum(1 for emp in sorted_employees if emp['has_vacation'])
    missing_employees = total_employees - active_employees - vacation_employees
    total_reports = sum(emp['reports_count'] for emp in sorted_employees)
    total_stores = sum(emp['stores_count'] for emp in sorted_employees)
    
    # Summary headers
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_header = ws.cell(row=summary_row, column=1)
    summary_header.value = "General Statistics"
    summary_header.font = Font(name='Calibri', size=12, bold=True, color='000000')
    summary_header.alignment = center_alignment
    
    # Summary data
    summary_data = [
        ('Total Employees:', total_employees),
        ('Active Employees:', active_employees),
        ('On Vacation:', vacation_employees),
        ('Missing Employees:', missing_employees),
        ('Total Reports:', total_reports),
        ('Total Stores Reported:', total_stores)
    ]
    
    for i, (label, value) in enumerate(summary_data):
        label_cell = ws.cell(row=summary_row + 1 + i, column=1)
        label_cell.value = label
        label_cell.font = Font(name='Calibri', size=10, bold=True)
        label_cell.alignment = left_alignment
        
        value_cell = ws.cell(row=summary_row + 1 + i, column=2)
        value_cell.value = value
        value_cell.font = data_font
        value_cell.alignment = center_alignment
        
        # Highlight different categories
        if label == 'On Vacation:' and value > 0:
            value_cell.font = Font(name='Calibri', size=10, bold=True, color='FF8C00')
        elif label == 'Missing Employees:' and value > 0:
            value_cell.font = Font(name='Calibri', size=10, bold=True, color='CC0000')
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # ============================================================================
    # ADD VACATION DETAILS TABLE
    # ============================================================================
    
    # Add vacation details table below the summary statistics
    vacation_table_start_row = summary_row + len(summary_data) + 3
    
    # Get all vacations in the date range for all employees
    from datetime import timedelta
    
    all_vacations = Vacation.query.filter(
        Vacation.vacation_date >= vacation_start_date,
        Vacation.vacation_date <= vacation_end_date
    ).order_by(Vacation.user_id, Vacation.vacation_date).all()
    
    # Group vacations by employee
    employee_vacations = {}
    for vacation in all_vacations:
        user = User.query.get(vacation.user_id)
        if user and not user.is_admin:
            if user.employee_code not in employee_vacations:
                employee_vacations[user.employee_code] = {
                    'name': user.employee_name,
                    'code': user.employee_code,
                    'vacation_dates': []
                }
            employee_vacations[user.employee_code]['vacation_dates'].append(vacation.vacation_date)
    
    # Only create vacation table if there are vacations
    if employee_vacations:
        # Title for vacation table
        ws.merge_cells(f'A{vacation_table_start_row}:F{vacation_table_start_row}')
        vacation_title_cell = ws.cell(row=vacation_table_start_row, column=1)
        vacation_title_cell.value = f"Vacation Details - From {start_date or 'Beginning'} To {end_date or 'End'}"
        vacation_title_cell.font = Font(name='Calibri', size=14, bold=True, color='000000')
        vacation_title_cell.alignment = center_alignment
        
        # Calculate all dates in the range
        date_range = []
        current_date = vacation_start_date
        while current_date <= vacation_end_date:
            date_range.append(current_date)
            current_date += timedelta(days=1)
        
        # Create headers for vacation table
        vacation_header_row = vacation_table_start_row + 2
        
        # First two columns: Employee Code and Name
        code_cell = ws.cell(row=vacation_header_row, column=1)
        code_cell.value = 'Employee Code'
        code_cell.font = header_font
        code_cell.fill = header_color
        code_cell.alignment = center_alignment
        code_cell.border = thin_border
        
        name_cell = ws.cell(row=vacation_header_row, column=2)
        name_cell.value = 'Employee Name'
        name_cell.font = header_font
        name_cell.fill = header_color
        name_cell.alignment = center_alignment
        name_cell.border = thin_border
        
        # Date columns
        for col_idx, vacation_date in enumerate(date_range, 3):
            date_cell = ws.cell(row=vacation_header_row, column=col_idx)
            date_cell.value = vacation_date.strftime('%Y-%m-%d')
            date_cell.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
            date_cell.fill = header_color
            date_cell.alignment = center_alignment
            date_cell.border = thin_border
            # Set narrow column width for dates
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # Sort employees by name
        sorted_vacation_employees = sorted(employee_vacations.values(), key=lambda x: x['name'])
        
        # Write vacation data
        vacation_data_row = vacation_header_row + 1
        for emp in sorted_vacation_employees:
            # Employee code
            code_cell = ws.cell(row=vacation_data_row, column=1)
            code_cell.value = emp['code']
            code_cell.border = thin_border
            code_cell.alignment = center_alignment
            code_cell.font = data_font
            
            # Employee name
            name_cell = ws.cell(row=vacation_data_row, column=2)
            name_cell.value = emp['name']
            name_cell.border = thin_border
            name_cell.alignment = left_alignment
            name_cell.font = data_font
            
            # Mark vacation dates
            for col_idx, check_date in enumerate(date_range, 3):
                date_cell = ws.cell(row=vacation_data_row, column=col_idx)
                
                if check_date in emp['vacation_dates']:
                    date_cell.value = '✓'
                    date_cell.fill = vacation_color
                    date_cell.font = Font(name='Calibri', size=12, bold=True, color='FF8C00')
                else:
                    date_cell.value = ''
                
                date_cell.border = thin_border
                date_cell.alignment = center_alignment
            
            vacation_data_row += 1
        
        # Add legend below vacation table
        legend_row = vacation_data_row + 2
        
        legend_cell = ws.cell(row=legend_row, column=1)
        legend_cell.value = 'Legend:'
        legend_cell.font = Font(name='Calibri', size=10, bold=True)
        
        legend_vacation_cell = ws.cell(row=legend_row + 1, column=1)
        legend_vacation_cell.value = '✓ = On Vacation'
        legend_vacation_cell.fill = vacation_color
        legend_vacation_cell.font = vacation_font
        legend_vacation_cell.border = thin_border
        legend_vacation_cell.alignment = center_alignment
        
        print(f"✅ Added vacation details table with {len(sorted_vacation_employees)} employees and {len(date_range)} days")
    else:
        print(f"ℹ️ No vacations found in the date range, skipping vacation details table")
    
    return ws


def _create_sheet_name(spvr_code, spvr_name, is_multi_spvr, used_names):
    """Create a clean, valid Excel sheet name with Name_Code format (no SPVR prefix)"""
    import re
    
    if not spvr_name or spvr_name.strip() == '':
        # Fallback to code only if name is missing
        print(f"Warning: SPVR name missing for code {spvr_code}, using code only")
        base_name = spvr_code if is_multi_spvr else "Report"
    else:
        # Clean the name: remove invalid characters and replace spaces with underscores
        clean_name = re.sub(r'[\\/*\[\]:?]', '', spvr_name.strip())
        clean_name = re.sub(r'\s+', '_', clean_name)  # Replace spaces with underscores
        
        # Clean the code: remove invalid characters
        clean_code = re.sub(r'[\\/*\[\]:?]', '', str(spvr_code).strip())
        
        if is_multi_spvr:
            # New format: <Name>_<Code> (no SPVR prefix)
            base_name = f"{clean_name}_{clean_code}"
        else:
            # Single SPVR: use name only or "Report" if name too long
            if len(clean_name) <= 25:  # Leave room for potential suffix
                base_name = clean_name
            else:
                base_name = "Report"
    
    # Ensure max 31 characters (Excel limit)
    if len(base_name) > 31:
        if is_multi_spvr:
            # Calculate available space for name
            clean_code = re.sub(r'[\\/*\[\]:?]', '', str(spvr_code).strip())
            code_part = f"_{clean_code}"
            available_for_name = 31 - len(code_part)
            
            if available_for_name > 3:  # Need at least 3 chars for name
                clean_name = re.sub(r'[\\/*\[\]:?]', '', spvr_name.strip())
                clean_name = re.sub(r'\s+', '_', clean_name)
                truncated_name = clean_name[:available_for_name]
                base_name = f"{truncated_name}{code_part}"
            else:
                # If code is too long, use code only
                base_name = clean_code[:31]
        else:
            # Single SPVR: truncate name
            base_name = base_name[:31]
    
    # Handle duplicates
    final_name = base_name
    counter = 1
    while final_name in used_names:
        suffix = f"_{counter}"
        if len(base_name) + len(suffix) <= 31:
            final_name = f"{base_name}{suffix}"
        else:
            # Truncate base name to make room for suffix
            truncated_base = base_name[:31-len(suffix)]
            final_name = f"{truncated_base}{suffix}"
        counter += 1
    
    return final_name


def _format_excel_sheet_enhanced(ws, reports, spvr_name, spvr_code):
    """Apply enhanced professional formatting matching the exact specified layout"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define exact colors as specified
    am_spvr_color = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')      # #E2EFDA - AM-SPVR & Store Information
    sales_color = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')       # #D9E1F2 - Sales Movement
    product_color = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')     # #FCE4D6 - Samsung Product Availability
    activities_color = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # #FFF2CC - Store Activities
    vod_result_color = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # #D9E1F2 - VOD & Result & Action
    bright_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')     # #FFFF00 - Bright yellow for specific cells
    
    # Fonts
    main_header_font = Font(name='Calibri', size=11, bold=True, color='000000')  # Larger for main headers
    header_font = Font(name='Calibri', size=10, bold=True, color='000000')
    data_font = Font(name='Calibri', size=9, bold=False, color='000000')
    
    # Alignment
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column definitions matching the exact specified structure
    columns = [
        # AM-SPVR & Store Information
        {'header': 'Date', 'width': 12},
        {'header': 'SPVR Code', 'width': 12},
        {'header': 'SPVR Name', 'width': 20},
        {'header': 'Shop code', 'width': 12},
        {'header': 'Shop Name', 'width': 25},
        {'header': 'Area', 'width': 15},
        {'header': 'Governorate', 'width': 15},
        
        # Sales Movement
        {'header': 'Samsung', 'width': 35},
        {'header': 'Competitors', 'width': 35},
        
        # Samsung Product Availability
        {'header': 'TV', 'width': 30},
        {'header': 'HA', 'width': 30},
        
        # Store Activities
        {'header': 'SFO, PMT', 'width': 25},
        {'header': 'Display', 'width': 25},
        {'header': 'Store Issue', 'width': 25},
        
        # VOD & Result & Action
        {'header': 'Complaints, Issues, Requirements', 'width': 40},
        {'header': 'Store, Member', 'width': 40}
    ]
    
    # Define the exact structure as specified
    # Row 1: Main Headers
    main_headers = [
        {'name': 'AM-SPVR & Store Information', 'start': 1, 'end': 7, 'color': am_spvr_color},
        {'name': 'Sales Movement', 'start': 8, 'end': 9, 'color': sales_color},
        {'name': 'Samsung Product Availability', 'start': 10, 'end': 11, 'color': product_color},
        {'name': 'Store Activities', 'start': 12, 'end': 14, 'color': activities_color},
        {'name': 'VOD', 'start': 15, 'end': 15, 'color': vod_result_color},
        {'name': 'Result & Action', 'start': 16, 'end': 16, 'color': vod_result_color}
    ]
    
    # Row 2: Sub Headers
    sub_headers = [
        {'name': 'Member Data', 'start': 1, 'end': 3, 'color': am_spvr_color},
        {'name': 'Shop Data', 'start': 4, 'end': 7, 'color': am_spvr_color},
        {'name': 'Samsung & Competitors (LG, Araby, Others)', 'start': 8, 'end': 9, 'color': sales_color},
        {'name': 'Ditributor, Key Model, Flag', 'start': 10, 'end': 10, 'color': product_color},
        {'name': 'Ditributor, Key Model, Flag', 'start': 11, 'end': 11, 'color': product_color},
        {'name': 'Samsung & Competitors', 'start': 12, 'end': 14, 'color': activities_color},
        {'name': 'Store & Dealer\'s Situation', 'start': 15, 'end': 15, 'color': bright_yellow},
        {'name': 'What I did ?', 'start': 16, 'end': 16, 'color': bright_yellow}
    ]
    
    # Write and format main headers (Row 1)
    for header in main_headers:
        # Merge cells for main header
        if header['start'] == header['end']:
            cell = ws.cell(row=1, column=header['start'])
        else:
            ws.merge_cells(start_row=1, start_column=header['start'], end_row=1, end_column=header['end'])
            cell = ws.cell(row=1, column=header['start'])
        
        cell.value = header['name']
        cell.font = main_header_font  # Use larger font for main headers
        cell.fill = header['color']
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Apply formatting to all merged cells
        for col in range(header['start'], header['end'] + 1):
            merged_cell = ws.cell(row=1, column=col)
            merged_cell.border = thin_border
            merged_cell.fill = header['color']
            merged_cell.font = main_header_font  # Use larger font for main headers
    
    # Write and format sub headers (Row 2)
    for sub_header in sub_headers:
        # Merge cells for sub header
        if sub_header['start'] == sub_header['end']:
            cell = ws.cell(row=2, column=sub_header['start'])
        else:
            ws.merge_cells(start_row=2, start_column=sub_header['start'], end_row=2, end_column=sub_header['end'])
            cell = ws.cell(row=2, column=sub_header['start'])
        
        cell.value = sub_header['name']
        cell.font = header_font
        cell.fill = sub_header['color']
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Apply formatting to all merged cells
        for col in range(sub_header['start'], sub_header['end'] + 1):
            merged_cell = ws.cell(row=2, column=col)
            merged_cell.border = thin_border
            merged_cell.fill = sub_header['color']
            merged_cell.font = header_font
    
    # Write and format column headers (Row 3)
    column_colors = [
        am_spvr_color, am_spvr_color, am_spvr_color, am_spvr_color, am_spvr_color, am_spvr_color, am_spvr_color,  # AM-SPVR & Store Information
        sales_color, sales_color,  # Sales Movement
        product_color, product_color,  # Samsung Product Availability
        activities_color, activities_color, activities_color,  # Store Activities (removed Sales)
        bright_yellow,  # VOD - Combined Complaints, Issues, Requirements
        bright_yellow   # Result & Action - Combined Store, Member
    ]
    
    for col_idx, col_info in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = col_info['header']
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = column_colors[col_idx - 1] if col_idx - 1 < len(column_colors) else am_spvr_color
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = col_info['width']
    
    # Write data rows
    for row_idx, report in enumerate(reports, 4):  # Start from row 4
        # Get governorate from store, area, or user's branch assignments
        governorate = ''
        
        # First try to get from store
        if hasattr(report, 'store') and hasattr(report.store, 'governorate') and report.store.governorate:
            governorate = report.store.governorate
        # Then try to get from area
        elif hasattr(report.area, 'governorate') and report.area.governorate:
            governorate = report.area.governorate
        # Finally try to get from user's branch assignments
        else:
            # Get the first branch with a governorate for this user
            user_branches = Branch.query.filter_by(owner_user_id=report.user_id).all()
            for branch in user_branches:
                if branch.governorate:
                    governorate = branch.governorate
                    break
        
        # Data mapping with Egypt local time
        report_date_local = utc_to_egypt_time(report.report_date)
        
        row_data = [
            report_date_local.strftime('%Y-%m-%d') if report_date_local else '',
            report.employee.employee_code,
            report.employee.employee_name,
            report.store.code,
            report.store.name,
            report.area.name,
            governorate,
            report.samsung_sales or '',
            report.competitors_sales or '',
            report.tv_availability or '',
            report.ha_availability or '',
            report.sfo_pmt or '',
            report.display_activities or '',
            report.store_issues or '',
            report.complaints or '',  # Combined complaints, issues, requirements
            report.actions_taken or ''  # Combined store, member
        ]
        
        # Write data to cells
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_alignment
    
    # Freeze panes (freeze first three rows)
    ws.freeze_panes = 'A4'
    
    # Set row heights
    ws.row_dimensions[1].height = 30  # Main headers - increased for better visibility
    ws.row_dimensions[2].height = 25  # Sub headers
    ws.row_dimensions[3].height = 25  # Column headers
    
    # Auto-adjust row heights for data rows based on content
    _auto_fit_rows(ws, reports, min_height=20, start_row=4)  # Start from row 4 (data rows)
    
    # Auto-adjust column widths based on content
    _auto_fit_columns(ws, reports, columns, start_row=4)
    
    # Add print settings
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    
    # Set print area to include all data
    max_row = len(reports) + 3  # 3 header rows + data rows
    max_col = len(columns)
    ws.print_area = f'A1:{get_column_letter(max_col)}{max_row}'
    
    return ws


def _auto_fit_rows(ws, reports, min_height=20, start_row=4):
    """Auto-fit row heights based on content with minimum height"""
    from openpyxl.utils import get_column_letter
    
    # Calculate row heights for data rows
    for row_idx in range(start_row, len(reports) + start_row):  # Start from specified row (data rows)
        max_lines = 1
        
        # Check each cell in the row to find the maximum number of lines
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                # Estimate number of lines based on content length and column width
                content = str(cell.value)
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 15
                
                # Estimate lines needed (rough calculation)
                chars_per_line = max(int(col_width * 0.8), 10)  # Conservative estimate
                estimated_lines = max(1, len(content) // chars_per_line + (1 if len(content) % chars_per_line > 0 else 0))
                
                # Account for explicit line breaks
                if '\n' in content:
                    explicit_lines = content.count('\n') + 1
                    estimated_lines = max(estimated_lines, explicit_lines)
                
                max_lines = max(max_lines, estimated_lines)
        
        # Set row height based on content (with minimum height)
        # Each line needs approximately 15 points, plus padding
        calculated_height = max(min_height, max_lines * 15 + 5)
        ws.row_dimensions[row_idx].height = min(calculated_height, 100)  # Cap at 100 points


def _auto_fit_columns(ws, reports, columns, start_row=4):
    """Auto-fit column widths based on content with margin"""
    from openpyxl.utils import get_column_letter
    
    # Calculate optimal width for each column
    for col_idx, col_info in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_width = len(col_info['header']) + 3  # Start with header width + margin
        
        # Check header content
        header_cell = ws.cell(row=3, column=col_idx)  # Column headers are now in row 3
        if header_cell.value:
            header_width = len(str(header_cell.value)) + 3
            max_width = max(max_width, header_width)
        
        # Check data content
        for row_idx in range(start_row, len(reports) + start_row):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                content = str(cell.value)
                
                # For multi-line content, use the longest line
                if '\n' in content:
                    lines = content.split('\n')
                    content_width = max(len(line) for line in lines) + 3
                else:
                    content_width = len(content) + 3
                
                max_width = max(max_width, content_width)
        
        # Apply calculated width with reasonable limits
        optimal_width = min(max_width, 50)  # Cap at 50 characters
        
        # Use the predefined width from columns config as minimum
        predefined_width = col_info.get('width', 10)
        optimal_width = max(optimal_width, predefined_width)  # Use predefined width as minimum
        
        ws.column_dimensions[col_letter].width = optimal_width

# ============================================================================
# COMMENTS & NOTIFICATIONS SYSTEM
# ============================================================================

@bp.route('/api/reports/<int:report_id>/comments', methods=['GET'])
@admin_required
def api_get_report_comments(report_id):
    """Get all comments for a report"""
    try:
        from app.models import ReportComment
        
        report = Report.query.get_or_404(report_id)
        comments = ReportComment.query.filter_by(report_id=report_id).order_by(ReportComment.created_at.desc()).all()
        
        comments_data = []
        for comment in comments:
            comment_time_egypt = utc_to_egypt_time(comment.created_at)
            comments_data.append({
                'id': comment.id,
                'comment_text': comment.comment_text,
                'commenter_name': comment.commenter.employee_name,
                'is_read': comment.is_read,
                'created_at': comment_time_egypt.strftime('%Y-%m-%d %H:%M') if comment_time_egypt else ''
            })
        
        return jsonify({
            'success': True,
            'comments': comments_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/reports/<int:report_id>/comments', methods=['POST'])
@admin_required
def api_add_report_comment(report_id):
    """Add a comment to a report"""
    try:
        from app.models import ReportComment, Notification
        
        report = Report.query.get_or_404(report_id)
        data = request.get_json()
        
        comment_text = data.get('comment_text', '').strip()
        
        # Allow empty comment if only changing status
        if not comment_text and not data.get('status'):
            return jsonify({'success': False, 'message': 'Comment text or status change is required'}), 400
        
        # Create comment only if text is provided
        comment = None
        if comment_text:
            comment = ReportComment(
                report_id=report_id,
                user_id=session['user_id'],
                comment_text=comment_text
            )
            db.session.add(comment)
            
            # Create notification for new comment
            admin_user = User.query.get(session['user_id'])
            notification = Notification(
                user_id=report.user_id,
                title='تعليق جديد على تقريرك',
                message=f'أضاف {admin_user.employee_name} تعليقاً على تقريرك',
                notification_type='new_comment',
                related_report_id=report_id
            )
            db.session.add(notification)
        
        # Update report status if provided
        new_status = data.get('status')
        if new_status and new_status in ['new', 'under_review', 'reviewed', 'needs_revision']:
            old_status = report.status
            report.status = new_status
            
            # Create notification for status change
            if old_status != new_status:
                status_names = {
                    'new': 'جديد',
                    'under_review': 'تحت المراجعة',
                    'reviewed': 'تمت المراجعة',
                    'needs_revision': 'يحتاج تعديل'
                }
                
                notification = Notification(
                    user_id=report.user_id,
                    title='تغيير حالة التقرير',
                    message=f'تم تغيير حالة تقريرك من "{status_names.get(old_status, old_status)}" إلى "{status_names.get(new_status, new_status)}"',
                    notification_type='status_change',
                    related_report_id=report_id
                )
                db.session.add(notification)
        
        db.session.commit()
        
        response_data = {
            'success': True,
            'message': 'Updated successfully'
        }
        
        if comment:
            comment_time_egypt = utc_to_egypt_time(comment.created_at)
            response_data['comment'] = {
                'id': comment.id,
                'comment_text': comment.comment_text,
                'commenter_name': comment.commenter.employee_name,
                'created_at': comment_time_egypt.strftime('%Y-%m-%d %H:%M') if comment_time_egypt else ''
            }
        
        return jsonify(response_data), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/comments/<int:comment_id>', methods=['PUT'])
@admin_required
def api_update_comment(comment_id):
    """Update a comment"""
    try:
        from app.models import ReportComment
        
        comment = ReportComment.query.get_or_404(comment_id)
        
        # Check if current user is the comment owner
        if comment.user_id != session['user_id']:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403
        
        data = request.get_json()
        comment_text = data.get('comment_text', '').strip()
        
        if not comment_text:
            return jsonify({'success': False, 'message': 'Comment text is required'}), 400
        
        comment.comment_text = comment_text
        db.session.commit()
        
        comment_time_egypt = utc_to_egypt_time(comment.created_at)
        
        return jsonify({
            'success': True,
            'message': 'Comment updated successfully',
            'comment': {
                'id': comment.id,
                'comment_text': comment.comment_text,
                'commenter_name': comment.commenter.employee_name,
                'created_at': comment_time_egypt.strftime('%Y-%m-%d %H:%M') if comment_time_egypt else ''
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/comments/<int:comment_id>', methods=['DELETE'])
@admin_required
def api_delete_comment(comment_id):
    """Delete a comment"""
    try:
        from app.models import ReportComment
        
        comment = ReportComment.query.get_or_404(comment_id)
        
        # Check if current user is the comment owner
        if comment.user_id != session['user_id']:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403
        
        db.session.delete(comment)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Comment deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/reports/<int:report_id>/status', methods=['PUT'])
@admin_required
def api_update_report_status(report_id):
    """Update report status and mark as read"""
    try:
        from app.models import Notification
        
        report = Report.query.get_or_404(report_id)
        data = request.get_json()
        
        new_status = data.get('status')
        mark_as_read = data.get('mark_as_read', False)
        
        if new_status and new_status in ['new', 'under_review', 'reviewed', 'needs_revision']:
            old_status = report.status
            report.status = new_status
            
            # Create notification for status change
            if old_status != new_status:
                status_names = {
                    'new': 'جديد',
                    'under_review': 'تحت المراجعة',
                    'reviewed': 'تمت المراجعة',
                    'needs_revision': 'يحتاج تعديل'
                }
                
                notification = Notification(
                    user_id=report.user_id,
                    title='تغيير حالة التقرير',
                    message=f'تم تغيير حالة تقريرك من "{status_names.get(old_status, old_status)}" إلى "{status_names.get(new_status, new_status)}"',
                    notification_type='status_change',
                    related_report_id=report_id
                )
                db.session.add(notification)
        
        if mark_as_read:
            report.is_read = True
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Report status updated successfully',
            'status': report.status,
            'is_read': report.is_read
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/notifications/unread-count', methods=['GET'])
@admin_required
def api_get_unread_notifications_count():
    """Get count of unread notifications for admin"""
    try:
        from app.models import Notification
        
        # For admin, count new reports as notifications
        new_reports_count = Report.query.filter_by(is_read=False).count()
        
        return jsonify({
            'success': True,
            'unread_count': new_reports_count
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/reports/stats', methods=['GET'])
@admin_required
def api_get_reports_stats():
    """Get reports statistics by status"""
    try:
        total_reports = Report.query.count()
        new_reports = Report.query.filter_by(status='new').count()
        under_review = Report.query.filter_by(status='under_review').count()
        reviewed = Report.query.filter_by(status='reviewed').count()
        needs_revision = Report.query.filter_by(status='needs_revision').count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total_reports,
                'new': new_reports,
                'under_review': under_review,
                'reviewed': reviewed,
                'needs_revision': needs_revision
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

from datetime import datetime, date
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz
from zoneinfo import ZoneInfo

# Timezone utilities for Egypt (Africa/Cairo) with automatic DST handling
def get_egypt_timezone():
    """Get Egypt timezone with automatic DST handling"""
    try:
        # Use zoneinfo (Python 3.9+) for better timezone handling
        return ZoneInfo("Africa/Cairo")
    except:
        # Fallback to pytz for older Python versions
        return pytz.timezone('Africa/Cairo')

def utc_to_egypt_time(utc_datetime):
    """Convert UTC datetime to Egypt local time with DST handling"""
    if utc_datetime is None:
        return None
    
    # Ensure the datetime is timezone-aware (UTC)
    if utc_datetime.tzinfo is None:
        utc_datetime = utc_datetime.replace(tzinfo=pytz.UTC)
    
    # Convert to Egypt timezone
    egypt_tz = get_egypt_timezone()
    return utc_datetime.astimezone(egypt_tz)

def egypt_time_to_utc(egypt_datetime):
    """Convert Egypt local time to UTC"""
    if egypt_datetime is None:
        return None
    
    egypt_tz = get_egypt_timezone()
    
    # If datetime is naive, assume it's in Egypt timezone
    if egypt_datetime.tzinfo is None:
        try:
            # Try zoneinfo first (Python 3.9+)
            egypt_datetime = egypt_datetime.replace(tzinfo=egypt_tz)
        except:
            # Fallback to pytz
            egypt_tz = pytz.timezone('Africa/Cairo')
            egypt_datetime = egypt_tz.localize(egypt_datetime)
    
    # Convert to UTC
    return egypt_datetime.astimezone(pytz.UTC)

def get_current_egypt_time():
    """Get current time in Egypt timezone"""
    utc_now = datetime.utcnow().replace(tzinfo=pytz.UTC)
    return utc_to_egypt_time(utc_now)

def parse_date_filter(date_str, is_end_date=False):
    """Parse date filter with proper timezone handling"""
    if not date_str:
        return None
    
    try:
        # Parse the date string (YYYY-MM-DD format from HTML date input)
        parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
        
        if is_end_date:
            # For end date, set to end of day (23:59:59) in Egypt timezone
            parsed_date = parsed_date.replace(hour=23, minute=59, second=59)
        else:
            # For start date, set to beginning of day (00:00:00) in Egypt timezone
            parsed_date = parsed_date.replace(hour=0, minute=0, second=0)
        
        # Convert Egypt local time to UTC for database query
        egypt_tz = get_egypt_timezone()
        
        try:
            # Try zoneinfo first (Python 3.9+)
            egypt_datetime = parsed_date.replace(tzinfo=egypt_tz)
        except:
            # Fallback to pytz
            egypt_tz = pytz.timezone('Africa/Cairo')
            egypt_datetime = egypt_tz.localize(parsed_date)
        
        # Convert to UTC and remove timezone info for database
        return egypt_datetime.astimezone(pytz.UTC).replace(tzinfo=None)
        
    except ValueError:
        return None

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or not session.get('is_admin', False):
            # Check if this is an API request
            if request.path.startswith('/admin/api/'):
                return jsonify({'success': False, 'error': 'Unauthorized'}), 401
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

@bp.route('/dashboard')
@admin_required
def dashboard():
    total_users = User.query.filter_by(is_admin=False).count()
    total_reports = Report.query.count()
    # Count unique branches by code (actual branches, not duplicates per supervisor)
    total_branches = db.session.query(db.func.count(db.func.distinct(Branch.code))).scalar() or 0
    # Count unique regions by name (actual regions, not duplicates per supervisor)
    total_regions = db.session.query(db.func.count(db.func.distinct(Region.name))).scalar() or 0
    
    # Recent reports with timezone conversion
    recent_reports_raw = Report.query.order_by(Report.created_at.desc()).limit(5).all()
    recent_reports = []
    for report in recent_reports_raw:
        # Convert UTC to Egypt time
        report.submitted_time_egypt = utc_to_egypt_time(report.created_at)
        recent_reports.append(report)
    
    return render_template('admin/dashboard.html', 
                         total_users=total_users,
                         total_reports=total_reports,
                         total_branches=total_branches,
                         total_regions=total_regions,
                         recent_reports=recent_reports,
                         utc_to_egypt_time=utc_to_egypt_time)

@bp.route('/api/dashboard/stats', methods=['GET'])
@admin_required
def get_dashboard_stats():
    """API endpoint to get real-time dashboard statistics"""
    try:
        # Get fresh counts from database using COUNT queries
        total_users = db.session.query(db.func.count(User.id)).filter(User.is_admin == False).scalar() or 0
        total_reports = db.session.query(db.func.count(Report.id)).scalar() or 0
        # Count unique branches by code (actual branches, not duplicates per supervisor)
        total_branches = db.session.query(db.func.count(db.func.distinct(Branch.code))).scalar() or 0
        # Count unique regions by name (actual regions, not duplicates per supervisor)
        total_regions = db.session.query(db.func.count(db.func.distinct(Region.name))).scalar() or 0
        
        return jsonify({
            'success': True,
            'stats': {
                'total_users': total_users,
                'total_reports': total_reports,
                'total_branches': total_branches,
                'total_regions': total_regions
            }
        })
    except Exception as e:
        print(f"Error in get_dashboard_stats: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================================
# NEW CLEAN USERS MANAGEMENT SYSTEM
# ============================================================================

@bp.route('/users')
@admin_required
def manage_users():
    """صفحة إدارة المستخدمين الجديدة والنظيفة"""
    return render_template('admin/users_management.html')

@bp.route('/admin-users')
@admin_required
def admin_users():
    """صفحة إدارة المستخدمين الإداريين"""
    return render_template('admin/admin_users_management.html')

# ============================================================================
# USERS API - Clean and Simple
# ============================================================================

@bp.route('/api/users', methods=['GET'])
@admin_required
def api_get_users():
    """Get all non-admin users with their regions and branches"""
    try:
        users = User.query.filter_by(is_admin=False).all()
        users_data = []
        
        for user in users:
            # Get user's owned regions and branches
            user_regions = Region.query.filter_by(owner_user_id=user.id).all()
            regions_data = []
            
            for region in user_regions:
                region_branches = Branch.query.filter_by(
                    owner_user_id=user.id, 
                    region_id=region.id
                ).all()
                
                branches_data = [{
                    'id': branch.id,
                    'name': branch.name,
                    'code': branch.code,
                    'governorate': branch.governorate
                } for branch in region_branches]
                
                regions_data.append({
                    'id': region.id,
                    'name': region.name,
                    'branches': branches_data
                })
            
            users_data.append({
                'id': user.id,
                'spvr_name': user.employee_name,
                'spvr_code': user.employee_code,
                'username': user.username,
                'created_at': user.created_at.isoformat(),
                'regions': regions_data
            })
        
        return jsonify(users_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    """Create a new user with regions and branches"""
    try:
        data = request.get_json()
        
        # Validate required fields
        spvr_name = data.get('spvr_name', '').strip()
        spvr_code = data.get('spvr_code', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        regions_data = data.get('regions', [])
        
        if not all([spvr_name, spvr_code, username, password]):
            return jsonify({
                'success': False, 
                'message': 'All required fields must be filled'
            }), 400
        
        # Check for existing username or employee code
        if User.query.filter_by(username=username).first():
            return jsonify({
                'success': False, 
                'message': 'Username already exists'
            }), 409
        
        if User.query.filter_by(employee_code=spvr_code).first():
            return jsonify({
                'success': False, 
                'message': 'SPVR Code already exists'
            }), 409
        
        # Create user
        user = User(
            employee_name=spvr_name,
            employee_code=spvr_code,
            username=username,
            password_hash=generate_password_hash(password),
            is_admin=False
        )
        
        db.session.add(user)
        db.session.flush()  # Get user ID
        
        # Create regions and branches
        created_regions = []
        created_branches = []
        
        for region_data in regions_data:
            region_name = region_data.get('name', '').strip()
            
            if not region_name:
                continue
            
            # Create region
            region = Region(
                name=region_name,
                owner_user_id=user.id
            )
            db.session.add(region)
            db.session.flush()  # Get region ID
            created_regions.append(region)
            
            # Create branches for this region
            branches_data = region_data.get('branches', [])
            for branch_data in branches_data:
                branch_name = branch_data.get('name', '').strip()
                branch_code = branch_data.get('code', '').strip()
                branch_governorate = branch_data.get('governorate', '').strip()
                
                if not branch_name or not branch_code:
                    continue
                
                # Check for duplicate branch codes for this user
                existing_branch = Branch.query.filter_by(
                    code=branch_code,
                    owner_user_id=user.id
                ).first()
                
                if existing_branch:
                    raise ValueError(f'Branch code "{branch_code}" already exists for this user')
                
                # Create branch
                branch = Branch(
                    name=branch_name,
                    code=branch_code,
                    governorate=branch_governorate if branch_governorate else None,
                    region_id=region.id,
                    owner_user_id=user.id
                )
                db.session.add(branch)
                created_branches.append(branch)
        
        # Commit all changes
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'User created successfully',
            'user_id': user.id,
            'regions_created': len(created_regions),
            'branches_created': len(created_branches)
        }), 201
        
    except ValueError as ve:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(ve)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users/<int:user_id>', methods=['GET'])
@admin_required
def api_get_user(user_id):
    """Get a specific user with regions and branches"""
    try:
        user = User.query.filter_by(id=user_id, is_admin=False).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get user's owned regions and branches
        user_regions = Region.query.filter_by(owner_user_id=user.id).all()
        regions_data = []
        
        for region in user_regions:
            region_branches = Branch.query.filter_by(
                owner_user_id=user.id, 
                region_id=region.id
            ).all()
            
            branches_data = [{
                'id': branch.id,
                'name': branch.name,
                'code': branch.code,
                'governorate': branch.governorate
            } for branch in region_branches]
            
            regions_data.append({
                'id': region.id,
                'name': region.name,
                'branches': branches_data
            })
        
        user_data = {
            'id': user.id,
            'spvr_name': user.employee_name,
            'spvr_code': user.employee_code,
            'username': user.username,
            'created_at': user.created_at.isoformat(),
            'regions': regions_data
        }
        
        return jsonify(user_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def api_update_user(user_id):
    """Update a user with regions and branches"""
    try:
        user = User.query.filter_by(id=user_id, is_admin=False).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        data = request.get_json()
        
        # Validate required fields
        spvr_name = data.get('spvr_name', '').strip()
        spvr_code = data.get('spvr_code', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        regions_data = data.get('regions', [])
        
        if not all([spvr_name, spvr_code, username]):
            return jsonify({
                'success': False, 
                'message': 'Name, code, and username are required'
            }), 400
        
        # Check for existing username or employee code (excluding current user)
        existing_user = User.query.filter(
            User.username == username,
            User.id != user_id
        ).first()
        if existing_user:
            return jsonify({
                'success': False, 
                'message': 'Username already exists'
            }), 409
        
        existing_user = User.query.filter(
            User.employee_code == spvr_code,
            User.id != user_id
        ).first()
        if existing_user:
            return jsonify({
                'success': False, 
                'message': 'SPVR Code already exists'
            }), 409
        
        # Update user basic info
        user.employee_name = spvr_name
        user.employee_code = spvr_code
        user.username = username
        
        if password:
            user.password_hash = generate_password_hash(password)
        
        # Delete existing regions and branches for this user
        existing_branches = Branch.query.filter_by(owner_user_id=user.id).all()
        for branch in existing_branches:
            db.session.delete(branch)
        
        existing_regions = Region.query.filter_by(owner_user_id=user.id).all()
        for region in existing_regions:
            db.session.delete(region)
        
        # Flush the deletions to ensure they're processed before creating new ones
        db.session.flush()
        
        # Create new regions and branches
        created_regions = []
        created_branches = []
        
        for region_data in regions_data:
            region_name = region_data.get('name', '').strip()
            
            if not region_name:
                continue
            
            # Create region
            region = Region(
                name=region_name,
                owner_user_id=user.id
            )
            db.session.add(region)
            db.session.flush()  # Get region ID
            created_regions.append(region)
            
            # Create branches for this region
            branches_data = region_data.get('branches', [])
            for branch_data in branches_data:
                branch_name = branch_data.get('name', '').strip()
                branch_code = branch_data.get('code', '').strip()
                branch_governorate = branch_data.get('governorate', '').strip()
                
                if not branch_name or not branch_code:
                    continue
                
                # Create branch
                branch = Branch(
                    name=branch_name,
                    code=branch_code,
                    governorate=branch_governorate if branch_governorate else None,
                    region_id=region.id,
                    owner_user_id=user.id
                )
                db.session.add(branch)
                created_branches.append(branch)
        
        # Commit all changes
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'User updated successfully',
            'regions_created': len(created_regions),
            'branches_created': len(created_branches)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_delete_user(user_id):
    """Delete a user and all their owned regions/branches"""
    try:
        user = User.query.filter_by(id=user_id, is_admin=False).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get request data to check for force delete or reassign options
        data = request.get_json(silent=True) or {}
        force_delete = data.get('force_delete', False)
        reassign_to_user_id = data.get('reassign_to_user_id')
        
        # Convert reassign_to_user_id to integer if it exists
        if reassign_to_user_id is not None:
            try:
                reassign_to_user_id = int(reassign_to_user_id)
            except (ValueError, TypeError):
                return jsonify({
                    'success': False, 
                    'message': 'Invalid user ID for reassignment'
                }), 400
        
        # Check if user has reports
        reports_count = len(user.reports)
        if reports_count > 0:
            if force_delete:
                # Delete all reports associated with this user
                for report in user.reports:
                    db.session.delete(report)
            else:
                # Return info about reports and force delete option
                return jsonify({
                    'success': False, 
                    'message': f'Cannot delete user. User has {reports_count} report(s). Use force delete to remove user and all reports.',
                    'reports_count': reports_count,
                    'force_delete_required': True
                }), 400
        
        # Delete user's branches first (due to foreign key constraints)
        user_branches = Branch.query.filter_by(owner_user_id=user.id).all()
        for branch in user_branches:
            db.session.delete(branch)
        
        # Delete user's regions
        user_regions = Region.query.filter_by(owner_user_id=user.id).all()
        for region in user_regions:
            db.session.delete(region)
        
        # Delete user's vacations (to avoid NOT NULL constraint error)
        from app.models import Vacation
        user_vacations = Vacation.query.filter_by(user_id=user.id).all()
        for vacation in user_vacations:
            db.session.delete(vacation)
        
        # Delete user's notifications (to avoid NOT NULL constraint error)
        user_notifications = Notification.query.filter_by(user_id=user.id).all()
        for notification in user_notifications:
            db.session.delete(notification)
        
        # Delete user's comments (to avoid NOT NULL constraint error)
        user_comments = ReportComment.query.filter_by(user_id=user.id).all()
        for comment in user_comments:
            db.session.delete(comment)
        
        # Delete user's audit logs (optional, but good for cleanup)
        user_audit_logs = AuditLog.query.filter_by(user_id=user.id).all()
        for audit_log in user_audit_logs:
            db.session.delete(audit_log)
        
        # Clear any associations (backward compatibility)
        user.assigned_areas.clear()
        user.assigned_stores.clear()
        user.assigned_regions.clear()
        user.assigned_branches.clear()
        
        # Delete user
        db.session.delete(user)
        db.session.commit()
        
        action_taken = "deleted"
        if reports_count > 0 and force_delete:
            action_taken = f"deleted along with {reports_count} reports"
        
        return jsonify({
            'success': True,
            'message': f'User {action_taken} successfully',
            'regions_deleted': len(user_regions),
            'branches_deleted': len(user_branches),
            'vacations_deleted': len(user_vacations),
            'reports_deleted': reports_count if force_delete else 0
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/users/<int:user_id>/reports', methods=['GET'])
@admin_required
def api_get_user_reports(user_id):
    """Get reports for a specific user"""
    try:
        user = User.query.filter_by(id=user_id, is_admin=False).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        reports_data = []
        for report in user.reports:
            # Convert UTC times to Egypt local time for display
            report_date_local = utc_to_egypt_time(report.report_date)
            created_at_local = utc_to_egypt_time(report.created_at)
            
            reports_data.append({
                'id': report.id,
                'store_name': report.store.name,
                'store_code': report.store.code,
                'area': report.area.name,
                'report_date': report_date_local.strftime('%Y-%m-%d') if report_date_local else '',
                'created_at': created_at_local.strftime('%Y-%m-%d %H:%M') if created_at_local else ''
            })
        
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'name': user.employee_name,
                'code': user.employee_code
            },
            'reports': reports_data,
            'reports_count': len(reports_data)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/reports/<int:report_id>', methods=['DELETE'])
@admin_required
def api_delete_report(report_id):
    """Delete a specific report"""
    try:
        report = Report.query.get(report_id)
        if not report:
            return jsonify({'success': False, 'message': 'Report not found'}), 404
        
        # Store report info for response
        report_info = {
            'id': report.id,
            'employee_name': report.employee.employee_name,
            'store_name': report.store.name,
            'report_date': report.report_date.strftime('%Y-%m-%d') if report.report_date else 'N/A'
        }
        
        # Delete the report
        db.session.delete(report)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Report deleted successfully',
            'deleted_report': report_info
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/reports/delete-all', methods=['DELETE'])
@admin_required
def api_delete_all_reports():
    """Delete all reports (with optional filters)"""
    try:
        # Get filter parameters (same as in api_reports)
        employee_name = request.args.get('employee_name', '')
        employee_code = request.args.get('employee_code', '')
        store_name = request.args.get('store_name', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        print(f"🗑️ Delete all reports request received")
        print(f"   Filters: employee_name={employee_name}, employee_code={employee_code}, store_name={store_name}")
        print(f"   Date range: {start_date} to {end_date}")
        
        # Build query with same filters as api_reports
        query = db.session.query(Report).join(User).join(Store).join(Area)
        
        if employee_name:
            query = query.filter(User.employee_name.contains(employee_name))
        if employee_code:
            query = query.filter(User.employee_code.contains(employee_code))
        if store_name:
            query = query.filter(Store.name.contains(store_name))
        
        # Date filtering
        if start_date:
            start_datetime = parse_date_filter(start_date, is_end_date=False)
            if start_datetime:
                query = query.filter(Report.report_date >= start_datetime)
        
        if end_date:
            end_datetime = parse_date_filter(end_date, is_end_date=True)
            if end_datetime:
                query = query.filter(Report.report_date <= end_datetime)
        
        # Get all matching reports
        reports_to_delete = query.all()
        deleted_count = len(reports_to_delete)
        
        print(f"   Found {deleted_count} reports to delete")
        
        if deleted_count == 0:
            return jsonify({
                'success': False,
                'message': 'No reports found matching the criteria',
                'deleted_count': 0
            }), 404
        
        # Delete all matching reports
        for report in reports_to_delete:
            db.session.delete(report)
        
        db.session.commit()
        
        print(f"✅ Successfully deleted {deleted_count} reports")
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted {deleted_count} report(s)',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        print(f"❌ Error deleting all reports: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================================
# ADMIN USERS MANAGEMENT API
# ============================================================================

@bp.route('/api/admin-users', methods=['GET'])
@admin_required
def api_get_admin_users():
    """Get all admin users"""
    try:
        admins = User.query.filter_by(is_admin=True).all()
        admins_data = []
        
        for admin in admins:
            # Determine if this is the primary admin (first admin created)
            is_primary = admin.id == 1
            
            admins_data.append({
                'id': admin.id,
                'employee_name': admin.employee_name,
                'employee_code': admin.employee_code,
                'username': admin.username,
                'is_primary_admin': is_primary,
                'created_at': admin.created_at.isoformat()
            })
        
        return jsonify(admins_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/admin-users', methods=['POST'])
@admin_required
def api_create_admin_user():
    """Create a new admin user"""
    try:
        data = request.get_json()
        
        # Validate required fields
        name = data.get('employee_name', '').strip()
        employee_code = data.get('employee_code', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not all([name, username, password]):
            return jsonify({
                'success': False, 
                'message': 'Name, username, and password are required'
            }), 400
        
        if len(password) < 8:
            return jsonify({
                'success': False, 
                'message': 'Password must be at least 8 characters'
            }), 400
        
        # Check for existing username
        if User.query.filter_by(username=username).first():
            return jsonify({
                'success': False, 
                'message': 'Username already exists'
            }), 409
        
        # Use provided employee code or generate one
        if not employee_code:
            import uuid
            employee_code = f"ADMIN_{uuid.uuid4().hex[:8].upper()}"
        
        # Create admin user
        admin = User(
            employee_name=name,
            employee_code=employee_code,
            username=username,
            password_hash=generate_password_hash(password),
            is_admin=True
        )
        
        db.session.add(admin)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Administrator created successfully',
            'admin_id': admin.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/admin-users/<int:admin_id>', methods=['PUT'])
@admin_required
def api_update_admin_user(admin_id):
    """Update an admin user"""
    try:
        admin = User.query.filter_by(id=admin_id, is_admin=True).first()
        if not admin:
            return jsonify({'success': False, 'message': 'Admin user not found'}), 404
        
        data = request.get_json()
        
        # Validate required fields
        name = data.get('employee_name', '').strip()
        employee_code = data.get('employee_code', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not all([name, username]):
            return jsonify({
                'success': False, 
                'message': 'Name and username are required'
            }), 400
        
        if password and len(password) < 8:
            return jsonify({
                'success': False, 
                'message': 'Password must be at least 8 characters'
            }), 400
        
        # Check for existing username (excluding current admin)
        existing_user = User.query.filter(
            User.username == username,
            User.id != admin_id
        ).first()
        if existing_user:
            return jsonify({
                'success': False, 
                'message': 'Username already exists'
            }), 409
        
        # Update admin
        admin.employee_name = name
        if employee_code:
            admin.employee_code = employee_code
        admin.username = username
        
        if password:
            admin.password_hash = generate_password_hash(password)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Administrator updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/admin-users/<int:admin_id>', methods=['DELETE'])
@admin_required
def api_delete_admin_user(admin_id):
    """Delete an admin user"""
    try:
        admin = User.query.filter_by(id=admin_id, is_admin=True).first()
        if not admin:
            return jsonify({'success': False, 'message': 'Admin user not found'}), 404
        
        # Check if this is the primary admin
        if admin.id == 1:
            return jsonify({
                'success': False, 
                'message': 'Cannot delete the primary administrator'
            }), 400
        
        # Check if this would be the last admin
        admin_count = User.query.filter_by(is_admin=True).count()
        if admin_count <= 1:
            return jsonify({
                'success': False, 
                'message': 'Cannot delete the last administrator account'
            }), 400
        
        # Delete admin's notifications (to avoid NOT NULL constraint error)
        admin_notifications = Notification.query.filter_by(user_id=admin.id).all()
        for notification in admin_notifications:
            db.session.delete(notification)
        
        # Delete admin's comments (to avoid NOT NULL constraint error)
        admin_comments = ReportComment.query.filter_by(user_id=admin.id).all()
        for comment in admin_comments:
            db.session.delete(comment)
        
        # Delete admin's audit logs (optional, but good for cleanup)
        admin_audit_logs = AuditLog.query.filter_by(user_id=admin.id).all()
        for audit_log in admin_audit_logs:
            db.session.delete(audit_log)
        
        # Delete admin
        db.session.delete(admin)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Administrator deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================================
# REPORTS MANAGEMENT (keeping existing functionality)
# ============================================================================

@bp.route('/reports')
@admin_required
def reports():
    return render_template('admin/reports.html')

@bp.route('/reports/<int:report_id>/view')
@admin_required
def view_report_detail(report_id):
    """عرض تفاصيل تقرير معين"""
    report = Report.query.get_or_404(report_id)
    
    # Mark report as read by admin
    if not report.is_read:
        report.is_read = True
        db.session.commit()
    
    return render_template('admin/view_report_detail.html', report=report)

@bp.route('/api/reports')
@admin_required
def api_reports():
    # Get filter parameters
    employee_name = request.args.get('employee_name', '')
    employee_code = request.args.get('employee_code', '')
    store_name = request.args.get('store_name', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build query
    query = db.session.query(Report).join(User).join(Store)
    
    if employee_name:
        query = query.filter(User.employee_name.contains(employee_name))
    if employee_code:
        query = query.filter(User.employee_code.contains(employee_code))
    if store_name:
        query = query.filter(Store.name.contains(store_name))
    
    # Fixed date filtering with proper timezone handling
    if start_date:
        start_datetime = parse_date_filter(start_date, is_end_date=False)
        if start_datetime:
            query = query.filter(Report.report_date >= start_datetime)
    
    if end_date:
        end_datetime = parse_date_filter(end_date, is_end_date=True)
        if end_datetime:
            query = query.filter(Report.report_date <= end_datetime)
    
    reports = query.order_by(Report.created_at.desc()).all()
    
    reports_data = []
    for report in reports:
        # Convert UTC times to Egypt local time for display
        report_date_local = utc_to_egypt_time(report.report_date)
        created_at_local = utc_to_egypt_time(report.created_at)
        
        # Count comments
        from app.models import ReportComment
        comments_count = ReportComment.query.filter_by(report_id=report.id).count()
        
        reports_data.append({
            'id': report.id,
            'employee_name': report.employee.employee_name,
            'employee_code': report.employee.employee_code,
            'store_name': report.store.name,
            'store_code': report.store.code,
            'area': report.area.name,
            'report_date': report_date_local.strftime('%Y-%m-%d %H:%M') if report_date_local else '',
            'created_at': created_at_local.strftime('%Y-%m-%d %H:%M') if created_at_local else '',
            'status': report.status if hasattr(report, 'status') else 'new',
            'is_read': report.is_read if hasattr(report, 'is_read') else False,
            'comments_count': comments_count
        })
    
    return jsonify(reports_data)

@bp.route('/preview_export')
@admin_required
def preview_export():
    """Preview export data before downloading"""
    # Get same filters as API
    employee_name = request.args.get('employee_name', '')
    employee_code = request.args.get('employee_code', '')
    store_name = request.args.get('store_name', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build query
    query = db.session.query(Report).join(User).join(Store).join(Area)
    
    if employee_name:
        query = query.filter(User.employee_name.contains(employee_name))
    if employee_code:
        query = query.filter(User.employee_code.contains(employee_code))
    if store_name:
        query = query.filter(Store.name.contains(store_name))
    
    # Fixed date filtering with proper timezone handling
    if start_date:
        start_datetime = parse_date_filter(start_date, is_end_date=False)
        if start_datetime:
            query = query.filter(Report.report_date >= start_datetime)
    
    if end_date:
        end_datetime = parse_date_filter(end_date, is_end_date=True)
        if end_datetime:
            query = query.filter(Report.report_date <= end_datetime)
    
    reports_raw = query.order_by(Report.created_at.desc()).limit(100).all()  # Limit for preview
    
    # Add governorate to each report
    reports = []
    for report in reports_raw:
        # Get governorate from store, area, or user's branch assignments
        governorate = ''
        
        # First try to get from store
        if hasattr(report, 'store') and hasattr(report.store, 'governorate') and report.store.governorate:
            governorate = report.store.governorate
        # Then try to get from area
        elif hasattr(report, 'area') and hasattr(report.area, 'governorate') and report.area.governorate:
            governorate = report.area.governorate
        # Finally try to get from user's branch assignments
        else:
            # Get the first branch with a governorate for this user
            user_branches = Branch.query.filter_by(owner_user_id=report.user_id).all()
            for branch in user_branches:
                if branch.governorate:
                    governorate = branch.governorate
                    break
        
        # Add governorate as attribute
        report.governorate_display = governorate
        reports.append(report)
    
    return render_template('admin/export_preview.html', reports=reports)

@bp.route('/export')
@admin_required
def export_reports():
    """Export reports with professional formatting and separate sheets per SPVR"""
    # Get same filters as API
    employee_name = request.args.get('employee_name', '')
    employee_code = request.args.get('employee_code', '')
    store_name = request.args.get('store_name', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build query
    query = db.session.query(Report).join(User).join(Store).join(Area)
    
    if employee_name:
        query = query.filter(User.employee_name.contains(employee_name))
    if employee_code:
        query = query.filter(User.employee_code.contains(employee_code))
    if store_name:
        query = query.filter(Store.name.contains(store_name))
    
    # Fixed date filtering with proper timezone handling
    if start_date:
        start_datetime = parse_date_filter(start_date, is_end_date=False)
        if start_datetime:
            query = query.filter(Report.report_date >= start_datetime)
    
    if end_date:
        end_datetime = parse_date_filter(end_date, is_end_date=True)
        if end_datetime:
            query = query.filter(Report.report_date <= end_datetime)
    
    reports = query.order_by(Report.created_at.desc()).all()
    
    # Generate filename with current Egypt local date and time
    current_egypt_time = get_current_egypt_time()
    filename = f'Report_{current_egypt_time.strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    if not reports:
        # Even if no reports, create summary sheet to show vacation status
        output = io.BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet with empty reports data
        summary_ws = wb.create_sheet(title="Reports Summary")
        _create_summary_sheet(summary_ws, {}, start_date, end_date)  # Empty spvr_reports dict
        
        # Create a "No Reports" sheet for information
        no_reports_ws = wb.create_sheet(title="No Reports Found")
        no_reports_ws.cell(row=1, column=1, value="No reports found for the selected criteria")
        no_reports_ws.cell(row=2, column=1, value=f"Date range: {start_date or 'All'} to {end_date or 'All'}")
        no_reports_ws.cell(row=3, column=1, value="Check the 'Reports Summary' sheet for employee vacation status")
        
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    # Group reports by SPVR
    spvr_reports = {}
    for report in reports:
        spvr_key = f"{report.employee.employee_code}_{report.employee.employee_name}"
        if spvr_key not in spvr_reports:
            spvr_reports[spvr_key] = []
        spvr_reports[spvr_key].append(report)
    
    # Create Excel file with professional formatting
    output = io.BytesIO()
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create summary sheet first
    summary_ws = wb.create_sheet(title="Reports Summary")
    _create_summary_sheet(summary_ws, spvr_reports, start_date, end_date)
    
    # Create a sheet for each SPVR with enhanced naming
    used_sheet_names = set()
    used_sheet_names.add("Reports Summary")  # Reserve summary sheet name
    
    for spvr_key, spvr_report_list in spvr_reports.items():
        spvr_code, spvr_name = spvr_key.split('_', 1)
        
        # Create enhanced sheet name with code and name
        sheet_name = _create_sheet_name(spvr_code, spvr_name, len(spvr_reports) > 1, used_sheet_names)
        used_sheet_names.add(sheet_name)
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Apply enhanced professional formatting to this sheet
        _format_excel_sheet_enhanced(ws, spvr_report_list, spvr_name, spvr_code)
    
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _create_summary_sheet(ws, spvr_reports, start_date, end_date):
    """Create summary sheet with employee statistics and missing employees in red"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models import User, Store, Branch, Vacation
    from datetime import datetime, date
    
    # Define colors and styles
    header_color = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue header
    summary_color = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Light green for active employees
    missing_color = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # Light red for missing employees
    vacation_color = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow for vacation
    
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10, bold=False, color='000000')
    missing_font = Font(name='Calibri', size=10, bold=True, color='CC0000')  # Red bold for missing
    vacation_font = Font(name='Calibri', size=10, bold=True, color='FF8C00')  # Orange bold for vacation
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Reports Summary - From {start_date or 'Beginning'} To {end_date or 'End'}"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='000000')
    title_cell.alignment = center_alignment
    
    # Headers
    headers = ['Employee Code', 'Employee Name', 'Stores Count', 'Reports Count', 'Last Report', 'Status']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_color
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Get all employees (non-admin users)
    all_employees = User.query.filter_by(is_admin=False).all()
    
    # Parse date range for vacation checking
    vacation_start_date = None
    vacation_end_date = None
    
    if start_date:
        try:
            vacation_start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        except:
            # If parsing fails, get earliest vacation date or use 30 days ago
            earliest_vacation = Vacation.query.order_by(Vacation.vacation_date.asc()).first()
            if earliest_vacation:
                vacation_start_date = earliest_vacation.vacation_date
            else:
                vacation_start_date = date.today() - timedelta(days=30)
    else:
        # If no start date provided, get earliest vacation date or use 30 days ago
        earliest_vacation = Vacation.query.order_by(Vacation.vacation_date.asc()).first()
        if earliest_vacation:
            vacation_start_date = earliest_vacation.vacation_date
        else:
            vacation_start_date = date.today() - timedelta(days=30)
        
    if end_date:
        try:
            vacation_end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except:
            # If parsing fails, get latest vacation date or use today
            latest_vacation = Vacation.query.order_by(Vacation.vacation_date.desc()).first()
            if latest_vacation:
                vacation_end_date = latest_vacation.vacation_date
            else:
                vacation_end_date = date.today()
    else:
        # If no end date provided, get latest vacation date or use today
        latest_vacation = Vacation.query.order_by(Vacation.vacation_date.desc()).first()
        if latest_vacation:
            vacation_end_date = latest_vacation.vacation_date
        else:
            vacation_end_date = date.today()
    
    print(f"🏖️ Checking vacations from {vacation_start_date} to {vacation_end_date}")
    
    # Create employee statistics
    employee_stats = {}
    
    # Process employees who submitted reports
    for spvr_key, reports in spvr_reports.items():
        spvr_code, spvr_name = spvr_key.split('_', 1)
        
        # Get unique stores reported by this employee
        unique_stores = set()
        last_report_date = None
        
        for report in reports:
            unique_stores.add(report.store.id)
            if not last_report_date or report.report_date > last_report_date:
                last_report_date = report.report_date
        
        employee_stats[spvr_code] = {
            'name': spvr_name,
            'code': spvr_code,
            'stores_count': len(unique_stores),
            'reports_count': len(reports),
            'last_report': last_report_date,
            'has_reports': True,
            'has_vacation': False
        }
    
    # Add employees who didn't submit reports
    for employee in all_employees:
        if employee.employee_code not in employee_stats:
            employee_stats[employee.employee_code] = {
                'name': employee.employee_name,
                'code': employee.employee_code,
                'stores_count': 0,
                'reports_count': 0,
                'last_report': None,
                'has_reports': False,
                'has_vacation': False
            }
    
    # Now check vacations for ALL employees (both with and without reports)
    for employee in all_employees:
        if employee.employee_code in employee_stats:
            # Check if employee has vacation in the date range - GET ALL vacations, not just first
            vacations = Vacation.query.filter(
                Vacation.user_id == employee.id,
                Vacation.vacation_date >= vacation_start_date,
                Vacation.vacation_date <= vacation_end_date
            ).all()
            
            has_vacation = len(vacations) > 0
            employee_stats[employee.employee_code]['has_vacation'] = has_vacation
            
            if has_vacation:
                vacation_dates = [v.vacation_date.strftime('%Y-%m-%d') for v in vacations]
                print(f"🏖️ Found {len(vacations)} vacation(s) for {employee.employee_name} ({employee.employee_code}): {', '.join(vacation_dates)}")
            else:
                print(f"💼 No vacation found for {employee.employee_name} ({employee.employee_code})")
    
    # Sort employees by name
    sorted_employees = sorted(employee_stats.values(), key=lambda x: x['name'])
    
    # Write data
    row_idx = 4
    for emp in sorted_employees:
        # Employee code
        cell = ws.cell(row=row_idx, column=1)
        cell.value = emp['code']
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Employee name
        cell = ws.cell(row=row_idx, column=2)
        cell.value = emp['name']
        cell.border = thin_border
        cell.alignment = left_alignment
        
        # Stores count
        cell = ws.cell(row=row_idx, column=3)
        cell.value = emp['stores_count']
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Reports count
        cell = ws.cell(row=row_idx, column=4)
        cell.value = emp['reports_count']
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Last report date
        cell = ws.cell(row=row_idx, column=5)
        if emp['last_report']:
            last_report_local = utc_to_egypt_time(emp['last_report'])
            cell.value = last_report_local.strftime('%Y-%m-%d') if last_report_local else ''
        else:
            cell.value = 'No Reports'
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Status
        cell = ws.cell(row=row_idx, column=6)
        if emp['has_vacation']:
            cell.value = 'On Vacation'
            cell.fill = vacation_color
            cell.font = vacation_font
        elif emp['has_reports']:
            cell.value = 'Active'
            cell.fill = summary_color
            cell.font = data_font
        else:
            cell.value = 'No Reports'
            cell.fill = missing_color
            cell.font = missing_font
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Apply row formatting
        for col in range(1, 7):
            row_cell = ws.cell(row=row_idx, column=col)
            if emp['has_vacation']:
                if col != 6:  # Don't override status cell
                    row_cell.fill = vacation_color
                    row_cell.font = vacation_font
            elif not emp['has_reports']:
                if col != 6:  # Don't override status cell
                    row_cell.fill = missing_color
                    row_cell.font = missing_font
            else:
                if col != 6:  # Don't override status cell
                    row_cell.font = data_font
        
        row_idx += 1
    
    # Set column widths
    column_widths = [15, 25, 15, 15, 15, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add summary statistics at the bottom
    summary_row = row_idx + 2
    
    total_employees = len(sorted_employees)
    active_employees = sum(1 for emp in sorted_employees if emp['has_reports'])
    vacation_employees = sum(1 for emp in sorted_employees if emp['has_vacation'])
    missing_employees = total_employees - active_employees - vacation_employees
    total_reports = sum(emp['reports_count'] for emp in sorted_employees)
    total_stores = sum(emp['stores_count'] for emp in sorted_employees)
    
    # Summary headers
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_header = ws.cell(row=summary_row, column=1)
    summary_header.value = "General Statistics"
    summary_header.font = Font(name='Calibri', size=12, bold=True, color='000000')
    summary_header.alignment = center_alignment
    
    # Summary data
    summary_data = [
        ('Total Employees:', total_employees),
        ('Active Employees:', active_employees),
        ('On Vacation:', vacation_employees),
        ('Missing Employees:', missing_employees),
        ('Total Reports:', total_reports),
        ('Total Stores Reported:', total_stores)
    ]
    
    for i, (label, value) in enumerate(summary_data):
        label_cell = ws.cell(row=summary_row + 1 + i, column=1)
        label_cell.value = label
        label_cell.font = Font(name='Calibri', size=10, bold=True)
        label_cell.alignment = left_alignment
        
        value_cell = ws.cell(row=summary_row + 1 + i, column=2)
        value_cell.value = value
        value_cell.font = data_font
        value_cell.alignment = center_alignment
        
        # Highlight different categories
        if label == 'On Vacation:' and value > 0:
            value_cell.font = Font(name='Calibri', size=10, bold=True, color='FF8C00')
        elif label == 'Missing Employees:' and value > 0:
            value_cell.font = Font(name='Calibri', size=10, bold=True, color='CC0000')
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # ============================================================================
    # ADD VACATION DETAILS TABLE
    # ============================================================================
    
    # Add vacation details table below the summary statistics
    vacation_table_start_row = summary_row + len(summary_data) + 3
    
    # Get all vacations in the date range for all employees
    from datetime import timedelta
    
    all_vacations = Vacation.query.filter(
        Vacation.vacation_date >= vacation_start_date,
        Vacation.vacation_date <= vacation_end_date
    ).order_by(Vacation.user_id, Vacation.vacation_date).all()
    
    # Group vacations by employee
    employee_vacations = {}
    for vacation in all_vacations:
        user = User.query.get(vacation.user_id)
        if user and not user.is_admin:
            if user.employee_code not in employee_vacations:
                employee_vacations[user.employee_code] = {
                    'name': user.employee_name,
                    'code': user.employee_code,
                    'vacation_dates': []
                }
            employee_vacations[user.employee_code]['vacation_dates'].append(vacation.vacation_date)
    
    # Only create vacation table if there are vacations
    if employee_vacations:
        # Title for vacation table
        ws.merge_cells(f'A{vacation_table_start_row}:F{vacation_table_start_row}')
        vacation_title_cell = ws.cell(row=vacation_table_start_row, column=1)
        vacation_title_cell.value = f"Vacation Details - From {start_date or 'Beginning'} To {end_date or 'End'}"
        vacation_title_cell.font = Font(name='Calibri', size=14, bold=True, color='000000')
        vacation_title_cell.alignment = center_alignment
        
        # Calculate all dates in the range
        date_range = []
        current_date = vacation_start_date
        while current_date <= vacation_end_date:
            date_range.append(current_date)
            current_date += timedelta(days=1)
        
        # Create headers for vacation table
        vacation_header_row = vacation_table_start_row + 2
        
        # First two columns: Employee Code and Name
        code_cell = ws.cell(row=vacation_header_row, column=1)
        code_cell.value = 'Employee Code'
        code_cell.font = header_font
        code_cell.fill = header_color
        code_cell.alignment = center_alignment
        code_cell.border = thin_border
        
        name_cell = ws.cell(row=vacation_header_row, column=2)
        name_cell.value = 'Employee Name'
        name_cell.font = header_font
        name_cell.fill = header_color
        name_cell.alignment = center_alignment
        name_cell.border = thin_border
        
        # Date columns
        for col_idx, vacation_date in enumerate(date_range, 3):
            date_cell = ws.cell(row=vacation_header_row, column=col_idx)
            date_cell.value = vacation_date.strftime('%Y-%m-%d')
            date_cell.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
            date_cell.fill = header_color
            date_cell.alignment = center_alignment
            date_cell.border = thin_border
            # Set narrow column width for dates
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # Sort employees by name
        sorted_vacation_employees = sorted(employee_vacations.values(), key=lambda x: x['name'])
        
        # Write vacation data
        vacation_data_row = vacation_header_row + 1
        for emp in sorted_vacation_employees:
            # Employee code
            code_cell = ws.cell(row=vacation_data_row, column=1)
            code_cell.value = emp['code']
            code_cell.border = thin_border
            code_cell.alignment = center_alignment
            code_cell.font = data_font
            
            # Employee name
            name_cell = ws.cell(row=vacation_data_row, column=2)
            name_cell.value = emp['name']
            name_cell.border = thin_border
            name_cell.alignment = left_alignment
            name_cell.font = data_font
            
            # Mark vacation dates
            for col_idx, check_date in enumerate(date_range, 3):
                date_cell = ws.cell(row=vacation_data_row, column=col_idx)
                
                if check_date in emp['vacation_dates']:
                    date_cell.value = '✓'
                    date_cell.fill = vacation_color
                    date_cell.font = Font(name='Calibri', size=12, bold=True, color='FF8C00')
                else:
                    date_cell.value = ''
                
                date_cell.border = thin_border
                date_cell.alignment = center_alignment
            
            vacation_data_row += 1
        
        # Add legend below vacation table
        legend_row = vacation_data_row + 2
        
        legend_cell = ws.cell(row=legend_row, column=1)
        legend_cell.value = 'Legend:'
        legend_cell.font = Font(name='Calibri', size=10, bold=True)
        
        legend_vacation_cell = ws.cell(row=legend_row + 1, column=1)
        legend_vacation_cell.value = '✓ = On Vacation'
        legend_vacation_cell.fill = vacation_color
        legend_vacation_cell.font = vacation_font
        legend_vacation_cell.border = thin_border
        legend_vacation_cell.alignment = center_alignment
        
        print(f"✅ Added vacation details table with {len(sorted_vacation_employees)} employees and {len(date_range)} days")
    else:
        print(f"ℹ️ No vacations found in the date range, skipping vacation details table")
    
    return ws


def _create_sheet_name(spvr_code, spvr_name, is_multi_spvr, used_names):
    """Create a clean, valid Excel sheet name with Name_Code format (no SPVR prefix)"""
    import re
    
    if not spvr_name or spvr_name.strip() == '':
        # Fallback to code only if name is missing
        print(f"Warning: SPVR name missing for code {spvr_code}, using code only")
        base_name = spvr_code if is_multi_spvr else "Report"
    else:
        # Clean the name: remove invalid characters and replace spaces with underscores
        clean_name = re.sub(r'[\\/*\[\]:?]', '', spvr_name.strip())
        clean_name = re.sub(r'\s+', '_', clean_name)  # Replace spaces with underscores
        
        # Clean the code: remove invalid characters
        clean_code = re.sub(r'[\\/*\[\]:?]', '', str(spvr_code).strip())
        
        if is_multi_spvr:
            # New format: <Name>_<Code> (no SPVR prefix)
            base_name = f"{clean_name}_{clean_code}"
        else:
            # Single SPVR: use name only or "Report" if name too long
            if len(clean_name) <= 25:  # Leave room for potential suffix
                base_name = clean_name
            else:
                base_name = "Report"
    
    # Ensure max 31 characters (Excel limit)
    if len(base_name) > 31:
        if is_multi_spvr:
            # Calculate available space for name
            clean_code = re.sub(r'[\\/*\[\]:?]', '', str(spvr_code).strip())
            code_part = f"_{clean_code}"
            available_for_name = 31 - len(code_part)
            
            if available_for_name > 3:  # Need at least 3 chars for name
                clean_name = re.sub(r'[\\/*\[\]:?]', '', spvr_name.strip())
                clean_name = re.sub(r'\s+', '_', clean_name)
                truncated_name = clean_name[:available_for_name]
                base_name = f"{truncated_name}{code_part}"
            else:
                # If code is too long, use code only
                base_name = clean_code[:31]
        else:
            # Single SPVR: truncate name
            base_name = base_name[:31]
    
    # Handle duplicates
    final_name = base_name
    counter = 1
    while final_name in used_names:
        suffix = f"_{counter}"
        if len(base_name) + len(suffix) <= 31:
            final_name = f"{base_name}{suffix}"
        else:
            # Truncate base name to make room for suffix
            truncated_base = base_name[:31-len(suffix)]
            final_name = f"{truncated_base}{suffix}"
        counter += 1
    
    return final_name


def _format_excel_sheet_enhanced(ws, reports, spvr_name, spvr_code):
    """Apply enhanced professional formatting matching the exact specified layout"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define exact colors as specified
    am_spvr_color = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')      # #E2EFDA - AM-SPVR & Store Information
    sales_color = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')       # #D9E1F2 - Sales Movement
    product_color = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')     # #FCE4D6 - Samsung Product Availability
    activities_color = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # #FFF2CC - Store Activities
    vod_result_color = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # #D9E1F2 - VOD & Result & Action
    bright_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')     # #FFFF00 - Bright yellow for specific cells
    
    # Fonts
    main_header_font = Font(name='Calibri', size=11, bold=True, color='000000')  # Larger for main headers
    header_font = Font(name='Calibri', size=10, bold=True, color='000000')
    data_font = Font(name='Calibri', size=9, bold=False, color='000000')
    
    # Alignment
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column definitions matching the exact specified structure
    columns = [
        # AM-SPVR & Store Information
        {'header': 'Date', 'width': 12},
        {'header': 'SPVR Code', 'width': 12},
        {'header': 'SPVR Name', 'width': 20},
        {'header': 'Shop code', 'width': 12},
        {'header': 'Shop Name', 'width': 25},
        {'header': 'Area', 'width': 15},
        {'header': 'Governorate', 'width': 15},
        
        # Sales Movement
        {'header': 'Samsung', 'width': 35},
        {'header': 'Competitors', 'width': 35},
        
        # Samsung Product Availability
        {'header': 'TV', 'width': 30},
        {'header': 'HA', 'width': 30},
        
        # Store Activities
        {'header': 'SFO, PMT', 'width': 25},
        {'header': 'Display', 'width': 25},
        {'header': 'Store Issue', 'width': 25},
        
        # VOD & Result & Action
        {'header': 'Complaints, Issues, Requirements', 'width': 40},
        {'header': 'Store, Member', 'width': 40}
    ]
    
    # Define the exact structure as specified
    # Row 1: Main Headers
    main_headers = [
        {'name': 'AM-SPVR & Store Information', 'start': 1, 'end': 7, 'color': am_spvr_color},
        {'name': 'Sales Movement', 'start': 8, 'end': 9, 'color': sales_color},
        {'name': 'Samsung Product Availability', 'start': 10, 'end': 11, 'color': product_color},
        {'name': 'Store Activities', 'start': 12, 'end': 14, 'color': activities_color},
        {'name': 'VOD', 'start': 15, 'end': 15, 'color': vod_result_color},
        {'name': 'Result & Action', 'start': 16, 'end': 16, 'color': vod_result_color}
    ]
    
    # Row 2: Sub Headers
    sub_headers = [
        {'name': 'Member Data', 'start': 1, 'end': 3, 'color': am_spvr_color},
        {'name': 'Shop Data', 'start': 4, 'end': 7, 'color': am_spvr_color},
        {'name': 'Samsung & Competitors (LG, Araby, Others)', 'start': 8, 'end': 9, 'color': sales_color},
        {'name': 'Ditributor, Key Model, Flag', 'start': 10, 'end': 10, 'color': product_color},
        {'name': 'Ditributor, Key Model, Flag', 'start': 11, 'end': 11, 'color': product_color},
        {'name': 'Samsung & Competitors', 'start': 12, 'end': 14, 'color': activities_color},
        {'name': 'Store & Dealer\'s Situation', 'start': 15, 'end': 15, 'color': bright_yellow},
        {'name': 'What I did ?', 'start': 16, 'end': 16, 'color': bright_yellow}
    ]
    
    # Write and format main headers (Row 1)
    for header in main_headers:
        # Merge cells for main header
        if header['start'] == header['end']:
            cell = ws.cell(row=1, column=header['start'])
        else:
            ws.merge_cells(start_row=1, start_column=header['start'], end_row=1, end_column=header['end'])
            cell = ws.cell(row=1, column=header['start'])
        
        cell.value = header['name']
        cell.font = main_header_font  # Use larger font for main headers
        cell.fill = header['color']
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Apply formatting to all merged cells
        for col in range(header['start'], header['end'] + 1):
            merged_cell = ws.cell(row=1, column=col)
            merged_cell.border = thin_border
            merged_cell.fill = header['color']
            merged_cell.font = main_header_font  # Use larger font for main headers
    
    # Write and format sub headers (Row 2)
    for sub_header in sub_headers:
        # Merge cells for sub header
        if sub_header['start'] == sub_header['end']:
            cell = ws.cell(row=2, column=sub_header['start'])
        else:
            ws.merge_cells(start_row=2, start_column=sub_header['start'], end_row=2, end_column=sub_header['end'])
            cell = ws.cell(row=2, column=sub_header['start'])
        
        cell.value = sub_header['name']
        cell.font = header_font
        cell.fill = sub_header['color']
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Apply formatting to all merged cells
        for col in range(sub_header['start'], sub_header['end'] + 1):
            merged_cell = ws.cell(row=2, column=col)
            merged_cell.border = thin_border
            merged_cell.fill = sub_header['color']
            merged_cell.font = header_font
    
    # Write and format column headers (Row 3)
    column_colors = [
        am_spvr_color, am_spvr_color, am_spvr_color, am_spvr_color, am_spvr_color, am_spvr_color, am_spvr_color,  # AM-SPVR & Store Information
        sales_color, sales_color,  # Sales Movement
        product_color, product_color,  # Samsung Product Availability
        activities_color, activities_color, activities_color,  # Store Activities (removed Sales)
        bright_yellow,  # VOD - Combined Complaints, Issues, Requirements
        bright_yellow   # Result & Action - Combined Store, Member
    ]
    
    for col_idx, col_info in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = col_info['header']
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = column_colors[col_idx - 1] if col_idx - 1 < len(column_colors) else am_spvr_color
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = col_info['width']
    
    # Write data rows
    for row_idx, report in enumerate(reports, 4):  # Start from row 4
        # Get governorate from store, area, or user's branch assignments
        governorate = ''
        
        # First try to get from store
        if hasattr(report, 'store') and hasattr(report.store, 'governorate') and report.store.governorate:
            governorate = report.store.governorate
        # Then try to get from area
        elif hasattr(report.area, 'governorate') and report.area.governorate:
            governorate = report.area.governorate
        # Finally try to get from user's branch assignments
        else:
            # Get the first branch with a governorate for this user
            user_branches = Branch.query.filter_by(owner_user_id=report.user_id).all()
            for branch in user_branches:
                if branch.governorate:
                    governorate = branch.governorate
                    break
        
        # Data mapping with Egypt local time
        report_date_local = utc_to_egypt_time(report.report_date)
        
        row_data = [
            report_date_local.strftime('%Y-%m-%d') if report_date_local else '',
            report.employee.employee_code,
            report.employee.employee_name,
            report.store.code,
            report.store.name,
            report.area.name,
            governorate,
            report.samsung_sales or '',
            report.competitors_sales or '',
            report.tv_availability or '',
            report.ha_availability or '',
            report.sfo_pmt or '',
            report.display_activities or '',
            report.store_issues or '',
            report.complaints or '',  # Combined complaints, issues, requirements
            report.actions_taken or ''  # Combined store, member
        ]
        
        # Write data to cells
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_alignment
    
    # Freeze panes (freeze first three rows)
    ws.freeze_panes = 'A4'
    
    # Set row heights
    ws.row_dimensions[1].height = 30  # Main headers - increased for better visibility
    ws.row_dimensions[2].height = 25  # Sub headers
    ws.row_dimensions[3].height = 25  # Column headers
    
    # Auto-adjust row heights for data rows based on content
    _auto_fit_rows(ws, reports, min_height=20, start_row=4)  # Start from row 4 (data rows)
    
    # Auto-adjust column widths based on content
    _auto_fit_columns(ws, reports, columns, start_row=4)
    
    # Add print settings
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    
    # Set print area to include all data
    max_row = len(reports) + 3  # 3 header rows + data rows
    max_col = len(columns)
    ws.print_area = f'A1:{get_column_letter(max_col)}{max_row}'
    
    return ws


def _auto_fit_rows(ws, reports, min_height=20, start_row=4):
    """Auto-fit row heights based on content with minimum height"""
    from openpyxl.utils import get_column_letter
    
    # Calculate row heights for data rows
    for row_idx in range(start_row, len(reports) + start_row):  # Start from specified row (data rows)
        max_lines = 1
        
        # Check each cell in the row to find the maximum number of lines
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                # Estimate number of lines based on content length and column width
                content = str(cell.value)
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 15
                
                # Estimate lines needed (rough calculation)
                chars_per_line = max(int(col_width * 0.8), 10)  # Conservative estimate
                estimated_lines = max(1, len(content) // chars_per_line + (1 if len(content) % chars_per_line > 0 else 0))
                
                # Account for explicit line breaks
                if '\n' in content:
                    explicit_lines = content.count('\n') + 1
                    estimated_lines = max(estimated_lines, explicit_lines)
                
                max_lines = max(max_lines, estimated_lines)
        
        # Set row height based on content (with minimum height)
        # Each line needs approximately 15 points, plus padding
        calculated_height = max(min_height, max_lines * 15 + 5)
        ws.row_dimensions[row_idx].height = min(calculated_height, 100)  # Cap at 100 points


def _auto_fit_columns(ws, reports, columns, start_row=4):
    """Auto-fit column widths based on content with margin"""
    from openpyxl.utils import get_column_letter
    
    # Calculate optimal width for each column
    for col_idx, col_info in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_width = len(col_info['header']) + 3  # Start with header width + margin
        
        # Check header content
        header_cell = ws.cell(row=3, column=col_idx)  # Column headers are now in row 3
        if header_cell.value:
            header_width = len(str(header_cell.value)) + 3
            max_width = max(max_width, header_width)
        
        # Check data content
        for row_idx in range(start_row, len(reports) + start_row):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                content = str(cell.value)
                
                # For multi-line content, use the longest line
                if '\n' in content:
                    lines = content.split('\n')
                    content_width = max(len(line) for line in lines) + 3
                else:
                    content_width = len(content) + 3
                
                max_width = max(max_width, content_width)
        
        # Apply calculated width with reasonable limits
        optimal_width = min(max_width, 50)  # Cap at 50 characters
        
        # Use the predefined width from columns config as minimum
        predefined_width = col_info.get('width', 10)
        optimal_width = max(optimal_width, predefined_width)  # Use predefined width as minimum
        
        ws.column_dimensions[col_letter].width = optimal_width

# ============================================================================
# COMMENTS & NOTIFICATIONS SYSTEM
# ============================================================================

@bp.route('/api/reports/<int:report_id>/comments', methods=['GET'])
@admin_required
def api_get_report_comments(report_id):
    """Get all comments for a report"""
    try:
        from app.models import ReportComment
        
        report = Report.query.get_or_404(report_id)
        comments = ReportComment.query.filter_by(report_id=report_id).order_by(ReportComment.created_at.desc()).all()
        
        comments_data = []
        for comment in comments:
            comment_time_egypt = utc_to_egypt_time(comment.created_at)
            comments_data.append({
                'id': comment.id,
                'comment_text': comment.comment_text,
                'commenter_name': comment.commenter.employee_name,
                'is_read': comment.is_read,
                'created_at': comment_time_egypt.strftime('%Y-%m-%d %H:%M') if comment_time_egypt else ''
            })
        
        return jsonify({
            'success': True,
            'comments': comments_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/reports/<int:report_id>/comments', methods=['POST'])
@admin_required
def api_add_report_comment(report_id):
    """Add a comment to a report"""
    try:
        from app.models import ReportComment, Notification
        
        report = Report.query.get_or_404(report_id)
        data = request.get_json()
        
        comment_text = data.get('comment_text', '').strip()
        
        # Allow empty comment if only changing status
        if not comment_text and not data.get('status'):
            return jsonify({'success': False, 'message': 'Comment text or status change is required'}), 400
        
        # Create comment only if text is provided
        comment = None
        if comment_text:
            comment = ReportComment(
                report_id=report_id,
                user_id=session['user_id'],
                comment_text=comment_text
            )
            db.session.add(comment)
            
            # Create notification for new comment
            admin_user = User.query.get(session['user_id'])
            notification = Notification(
                user_id=report.user_id,
                title='تعليق جديد على تقريرك',
                message=f'أضاف {admin_user.employee_name} تعليقاً على تقريرك',
                notification_type='new_comment',
                related_report_id=report_id
            )
            db.session.add(notification)
        
        # Update report status if provided
        new_status = data.get('status')
        if new_status and new_status in ['new', 'under_review', 'reviewed', 'needs_revision']:
            old_status = report.status
            report.status = new_status
            
            # Create notification for status change
            if old_status != new_status:
                status_names = {
                    'new': 'جديد',
                    'under_review': 'تحت المراجعة',
                    'reviewed': 'تمت المراجعة',
                    'needs_revision': 'يحتاج تعديل'
                }
                
                notification = Notification(
                    user_id=report.user_id,
                    title='تغيير حالة التقرير',
                    message=f'تم تغيير حالة تقريرك من "{status_names.get(old_status, old_status)}" إلى "{status_names.get(new_status, new_status)}"',
                    notification_type='status_change',
                    related_report_id=report_id
                )
                db.session.add(notification)
        
        db.session.commit()
        
        response_data = {
            'success': True,
            'message': 'Updated successfully'
        }
        
        if comment:
            comment_time_egypt = utc_to_egypt_time(comment.created_at)
            response_data['comment'] = {
                'id': comment.id,
                'comment_text': comment.comment_text,
                'commenter_name': comment.commenter.employee_name,
                'created_at': comment_time_egypt.strftime('%Y-%m-%d %H:%M') if comment_time_egypt else ''
            }
        
        return jsonify(response_data), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/comments/<int:comment_id>', methods=['PUT'])
@admin_required
def api_update_comment(comment_id):
    """Update a comment"""
    try:
        from app.models import ReportComment
        
        comment = ReportComment.query.get_or_404(comment_id)
        
